import sys
import os
import json
import random
import string
import smtplib
import ssl
import fitz  # PyMuPDF
import cv2
import numpy as np
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import io
import base64

# Suppress MuPDF error messages to keep the terminal clean
fitz.TOOLS.mupdf_display_errors(False)

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QPushButton, QFileDialog, QLabel, QProgressBar,
    QMessageBox, QFrame, QLineEdit, QStackedWidget, QHBoxLayout,
    QDialog, QFormLayout, QSplashScreen, QComboBox, QSpacerItem, QSizePolicy, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt5.QtGui import QIcon, QPixmap, QMovie, QFont, QColor
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize, QSettings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import google.generativeai as genai
import requests
from datetime import datetime
import time
import logging
import concurrent.futures
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# ---------------------------
# Global Constants and Colors
# ---------------------------
COLORS = {
    'primary': '#002175',
    'secondary': '#FF8A00',
    'tertiary': '#DAB284',
    'accent': '#DD6F20',
    'background': '#F6F6F6',
    'button': '#355FB8',
    'button_hover': '#487BEA',
    'progress': '#7EA3F1',
    'border': '#7384CB',
    'text': '#8497B0',
    'highlight': '#ECA86A',
    # Additional colors for login UI:
    'login_primary': '#002175',
    'login_secondary': '#FF8A00',
    'login_accent': '#284DAB',
    'login_alert': '#DD6F20',
    'login_bg': '#F6F6F6',
    'login_disabled': '#B9B9B9'
}

LOGO_URL = "https://marblebox.com/wp-content/uploads/2024/12/logo.svg"
ICON_URL = "https://hebbkx1anhila5yf.public.blob.vercel-storage.com/Marble%20Box%20logo%20Symbol_Colour-6MaHGSNNlEiGfa9Puy91URxaCtw5QA.png"

# ---------------------------
# SharePoint and Share Drive Configuration
# ---------------------------
SHAREPOINT_USERNAME = "sharepoint_user"
SHAREPOINT_PASSWORD = "sharepoint_password"
SHAREDRIVE_FOLDER = r"\\192.168.48.10\automation\SHAREPOINT DRIVE"
SHAREDRIVE_COMMERCIAL = os.path.join(SHAREDRIVE_FOLDER, "Commercial Line")
SHAREDRIVE_PERSONAL = os.path.join(SHAREDRIVE_FOLDER, "Personal Line")

# ---------------------------
# Enhanced Email Configuration
# ---------------------------
EMAIL_CONFIG = {
    'smtp_server': 'smtp.office365.com',
    'smtp_port': 587,
    'sender_email': 'noreply_automation@marblebox.com',
    'sender_password': 'M@rbleb0x@bfg642025$',
    'use_tls': True,
    'timeout': 30,
    'retry_attempts': 3,
    'retry_delay': 2
}

# ---------------------------
# Global variable for Email Threads
# ---------------------------
email_threads = []

# ---------------------------
# User Data File (for login)
# ---------------------------
USER_DATA_FILE = "users.json"
if not os.path.exists(USER_DATA_FILE):
    with open(USER_DATA_FILE, "w") as f:
        json.dump({}, f)

def load_users():
    with open(USER_DATA_FILE, "r") as f:
        return json.load(f)

def save_users(users):
    with open(USER_DATA_FILE, "w") as f:
        json.dump(users, f, indent=4)

def generate_temp_code(length=6):
    return ''.join(random.choices(string.digits, k=length))

# ---------------------------
# Enhanced Email Sending (in a separate thread)
# ---------------------------
class EmailThread(QThread):
    finished_signal = pyqtSignal(bool, str)  # success, error_message
    progress_signal = pyqtSignal(str)  # status message

    def __init__(self, recipient, subject, body, attachments=None, parent=None):
        super().__init__(parent)
        self.recipient = recipient
        self.subject = subject
        self.body = body
        self.attachments = attachments or []
        self.config = EMAIL_CONFIG

    def run(self):
        """Enhanced email sending with better error handling and retry logic"""
        for attempt in range(self.config['retry_attempts']):
            try:
                self.progress_signal.emit(f"Attempting to send email (attempt {attempt + 1}/{self.config['retry_attempts']})...")
                
                # Create message
                msg = MIMEMultipart()
                msg['From'] = self.config['sender_email']
                msg['To'] = self.recipient
                msg['Subject'] = self.subject
                
                # Add body
                msg.attach(MIMEText(self.body, 'plain'))
                
                # Add attachments if any
                for attachment_path in self.attachments:
                    if os.path.exists(attachment_path):
                        with open(attachment_path, "rb") as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                            encoders.encode_base64(part)
                            part.add_header(
                                'Content-Disposition',
                                f'attachment; filename= {os.path.basename(attachment_path)}'
                            )
                            msg.attach(part)
                
                # Create secure connection and send
                context = ssl.create_default_context()
                
                self.progress_signal.emit("Connecting to SMTP server...")
                with smtplib.SMTP(self.config['smtp_server'], self.config['smtp_port'], timeout=self.config['timeout']) as server:
                    server.ehlo()
                    
                    if self.config['use_tls']:
                        self.progress_signal.emit("Starting TLS encryption...")
                        server.starttls(context=context)
                        server.ehlo()
                    
                    self.progress_signal.emit("Authenticating...")
                    server.login(self.config['sender_email'], self.config['sender_password'])
                    
                    self.progress_signal.emit("Sending email...")
                    text = msg.as_string()
                    server.sendmail(self.config['sender_email'], self.recipient, text)
                    
                self.progress_signal.emit("Email sent successfully!")
                self.finished_signal.emit(True, "")
                return
                
            except smtplib.SMTPAuthenticationError as e:
                error_msg = f"SMTP Authentication failed: {str(e)}"
                logging.error(error_msg)
                if attempt == self.config['retry_attempts'] - 1:
                    self.finished_signal.emit(False, error_msg)
                    return
                    
            except smtplib.SMTPRecipientsRefused as e:
                error_msg = f"Recipient refused: {str(e)}"
                logging.error(error_msg)
                self.finished_signal.emit(False, error_msg)
                return
                
            except smtplib.SMTPServerDisconnected as e:
                error_msg = f"SMTP server disconnected: {str(e)}"
                logging.error(error_msg)
                if attempt == self.config['retry_attempts'] - 1:
                    self.finished_signal.emit(False, error_msg)
                    return
                    
            except Exception as e:
                error_msg = f"Email send error: {str(e)}"
                logging.error(error_msg)
                if attempt == self.config['retry_attempts'] - 1:
                    self.finished_signal.emit(False, error_msg)
                    return
            
            # Wait before retry
            if attempt < self.config['retry_attempts'] - 1:
                self.progress_signal.emit(f"Retrying in {self.config['retry_delay']} seconds...")
                time.sleep(self.config['retry_delay'])

def send_email(recipient, subject, body, attachments=None):
    """Enhanced email sending function with better error handling"""
    global email_threads
    thread = EmailThread(recipient, subject, body, attachments)
    
    def cleanup(success, error):
        if thread in email_threads:
            email_threads.remove(thread)
        if not success:
            logging.error(f"Email delivery failed: {error}")
    
    def progress_update(message):
        logging.info(f"Email progress: {message}")
    
    thread.finished_signal.connect(cleanup)
    thread.progress_signal.connect(progress_update)
    email_threads.append(thread)
    thread.start()
    return thread

# ---------------------------
# Enhanced OCR Processing Class
# ---------------------------
class EnhancedOCR:
    """Enhanced OCR class for better text extraction from challenging documents"""
    
    def __init__(self):
        # Configure Tesseract if available
        try:
            # Try to find tesseract executable
            if os.name == 'nt':  # Windows
                possible_paths = [
                    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
                    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
                    r'C:\Users\{}\AppData\Local\Tesseract-OCR\tesseract.exe'.format(os.getenv('USERNAME'))
                ]
                for path in possible_paths:
                    if os.path.exists(path):
                        pytesseract.pytesseract.tesseract_cmd = path
                        break
            
            # Test if tesseract is working
            pytesseract.get_tesseract_version()
            self.tesseract_available = True
            logging.info("Tesseract OCR is available and configured")
        except Exception as e:
            self.tesseract_available = False
            logging.warning(f"Tesseract OCR not available: {e}")
    
    def preprocess_image(self, image):
        """Apply various preprocessing techniques to improve OCR accuracy"""
        try:
            # Convert PIL image to OpenCV format
            if isinstance(image, Image.Image):
                opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            else:
                opencv_image = image
            
            # Convert to grayscale
            gray = cv2.cvtColor(opencv_image, cv2.COLOR_BGR2GRAY)
            
            # Apply different preprocessing techniques
            processed_images = []
            
            # Original grayscale
            processed_images.append(('original', gray))
            
            # Gaussian blur + threshold
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            _, thresh1 = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            processed_images.append(('gaussian_otsu', thresh1))
            
            # Adaptive threshold
            adaptive_thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            processed_images.append(('adaptive', adaptive_thresh))
            
            # Morphological operations
            kernel = np.ones((2, 2), np.uint8)
            morph = cv2.morphologyEx(thresh1, cv2.MORPH_CLOSE, kernel)
            processed_images.append(('morphological', morph))
            
            # Dilation and erosion
            dilated = cv2.dilate(thresh1, kernel, iterations=1)
            eroded = cv2.erode(dilated, kernel, iterations=1)
            processed_images.append(('dilate_erode', eroded))
            
            return processed_images
            
        except Exception as e:
            logging.error(f"Error in image preprocessing: {e}")
            return [('original', image)]
    
    def extract_text_with_ocr(self, image):
        """Extract text using multiple OCR techniques"""
        if not self.tesseract_available:
            return ""
        
        try:
            processed_images = self.preprocess_image(image)
            best_text = ""
            max_confidence = 0
            
            for name, processed_img in processed_images:
                try:
                    # Convert back to PIL Image for tesseract
                    pil_img = Image.fromarray(processed_img)
                    
                    # Try different PSM modes
                    psm_modes = [6, 8, 13, 3, 4]  # Different page segmentation modes
                    
                    for psm in psm_modes:
                        try:
                            custom_config = f'--oem 3 --psm {psm} -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,/$%()-: '
                            
                            # Get text with confidence
                            data = pytesseract.image_to_data(pil_img, config=custom_config, output_type=pytesseract.Output.DICT)
                            
                            # Calculate average confidence
                            confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
                            if confidences:
                                avg_confidence = sum(confidences) / len(confidences)
                                text = pytesseract.image_to_string(pil_img, config=custom_config).strip()
                                
                                if avg_confidence > max_confidence and len(text) > len(best_text):
                                    max_confidence = avg_confidence
                                    best_text = text
                                    logging.info(f"Better OCR result with {name} preprocessing, PSM {psm}, confidence: {avg_confidence:.2f}")
                        
                        except Exception as e:
                            continue
                
                except Exception as e:
                    logging.error(f"Error processing {name}: {e}")
                    continue
            
            return best_text
            
        except Exception as e:
            logging.error(f"Error in OCR text extraction: {e}")
            return ""
    
    def enhance_image_quality(self, image):
        """Enhance image quality for better OCR results"""
        try:
            if isinstance(image, np.ndarray):
                image = Image.fromarray(image)
            
            # Increase contrast
            enhancer = ImageEnhance.Contrast(image)
            image = enhancer.enhance(1.5)
            
            # Increase sharpness
            enhancer = ImageEnhance.Sharpness(image)
            image = enhancer.enhance(2.0)
            
            # Apply unsharp mask
            image = image.filter(ImageFilter.UnsharpMask(radius=2, percent=150, threshold=3))
            
            return image
            
        except Exception as e:
            logging.error(f"Error enhancing image quality: {e}")
            return image

# ---------------------------
# Field definitions (Default for Workers Compensation)
# ---------------------------
FIELDS_TO_EXTRACT = {
    "Policy Number": {
        "context": ["policy #", "policy number", "policy no", "policy num", "policy #:", "policy number:", "policy no:"],
        "format": "alphanumeric",
        "table_headers": ["policy", "policy number", "policy #"]
    },
    "Renewal of Policy": {
        "context": [
            "renewal", "renewal of", "prior policy", "previous policy", "renewing", "renewal policy",
            "renewal #", "renewal policy number", "renewing policy",
            "renewal coverage", "this coverage is renewed from",
            "policy renewal", "nonrenewal", "renewal effective date", "renewal date", "policy renewal date"
        ],
        "format": "alphanumeric",
        "table_headers": ["renewal", "renewal of", "prior policy"],
        "special_rules": {
            "look_near": ["policy number", "policy information"],
            "require_context": True
        }
    },
    "Policy Period From": {
        "context": ["period from", "effective date", "policy period is from", "effective", "beginning", "start date", "policy period:", "effective:", "policy effective date"],
        "format": "date",
        "table_headers": ["effective date", "period from", "start date"],
        "special_rules": {
            "date_format": True,
            "look_near": ["policy period", "effective date"]
        }
    },
    "Policy To": {
        "context": ["period to", "expiration date", "policy period is to", "expiration", "ending", "end date", "expires:", "expiration:", "policy expiration date"],
        "format": "date",
        "table_headers": ["expiration date", "period to", "end date"],
        "special_rules": {
            "date_format": True,
            "look_near": ["policy period", "expiration date"]
        }
    },
    "Coverage Provided By": {
        "context": ["coverage provided by", "insurance company", "carrier", "insurer", "insurance carrier", "underwritten by", "issued by", "writing company", "insurance provider"],
        "format": "text",
        "table_headers": ["carrier", "insurer", "company"]
    },
    "The Insured's Name": {
        "context": ["insured name", "named insured", "the insured", "insured:", "name of insured", "first named insured", "insured's name:", "named insured:"],
        "format": "text",
        "table_headers": ["insured name", "named insured"]
    },
    "DBA Name": {
        "context": ["dba", "doing business as", "trade name", "t/a", "trading as", "also known as", "aka", "d/b/a", "d.b.a.", "trading name", "business name", "operating as", "o/a", "registered as", "known as", "formerly known as", "fka"],
        "format": "text",
        "table_headers": ["dba", "trade name", "business name"],
        "special_rules": {
            "look_near": ["insured name", "business name"],
            "exclude_patterns": ["policy", "insurance", "company"],
            "require_context": True
        }
    },
    "Mailing Address": {
        "context": ["mailing address", "address", "location", "insured address", "business address", "address of insured", "principal address", "insured's address"],
        "format": "address",
        "table_headers": ["address", "mailing address", "location"]
    },
    "Type of Entity": {
        "context": ["type of entity", "business type", "organization type", "legal entity", "entity type", "form of business", "business organization", "legal structure"],
        "format": "text",
        "table_headers": ["entity type", "business type"]
    },
    "FEIN Number": {
        "context": ["fein", "tax id", "employer id", "federal id", "ein", "federal employer identification number", "employer identification number", "tax identification number"],
        "format": "numeric",
        "table_headers": ["fein", "tax id", "ein"]
    },
    "Total Estimated Annual Premium": {
        "context": ["total estimated annual premium", "estimated annual premium", "total premium", "estimated premium", "annual premium", "premium total", "total estimated premium"],
        "format": "currency",
        "table_headers": ["total premium", "estimated premium"],
        "special_rules": {
            "currency_format": True,
            "sum_if_multiple": True
        }
    },
    "Workers Compensation Insurance - 3.A": {
        "context": ["workers compensation insurance", "part one", "3.a", "part 3.a", "workers comp coverage", "part one - workers compensation", "workers compensation and employers liability insurance policy", "part one - workers compensation insurance", "statutory coverage", "workers compensation law", "compensation law coverage", "employee reliability insurance"],
        "format": "text",
        "table_headers": ["workers compensation", "part one", "coverage a"],
        "special_rules": {
            "look_for_states": True,
            "include_statutory": True
        }
    },
    "Employers Liability Insurance - 3.B": {
        "context": ["employers liability insurance", "part two", "3.b", "part 3.b", "employers liability coverage", "part two - employers liability", "coverage b", "employer's liability", "employers liability limits"],
        "format": "text",
        "table_headers": ["employers liability", "part two", "coverage b"],
        "special_rules": {
            "fixed_value": "WORKERS COMPENSATION AND EMPLOYERS LIABILITY INSURANCE POLICY"
        }
    },
    "Bodily injury by accident - Each Accident": {
        "context": ["bodily injury by accident", "each accident", "accident limit", "bi by accident", "bodily injury - each accident", "e.l. each accident"],
        "format": "currency",
        "table_headers": ["bodily injury", "each accident", "accident limit"],
        "special_rules": {
            "currency_format": True,
            "look_near": ["employers liability", "coverage b", "e.l."]
        }
    },
    "Bodily injury by disease - Policy Limit": {
        "context": ["bodily injury by disease", "policy limit", "disease - policy limit", "bi by disease - policy", "disease policy limit", "e.l. disease - policy limit"],
        "format": "currency",
        "table_headers": ["disease", "policy limit"],
        "special_rules": {
            "currency_format": True,
            "look_near": ["employers liability", "coverage b", "e.l."]
        }
    },
    "Bodily injury by disease - Each Employee": {
        "context": ["bodily injury by disease", "each employee", "disease - each employee", "bi by disease - employee", "disease per employee", "e.l. disease - each employee"],
        "format": "currency",
        "table_headers": ["disease", "each employee"],
        "special_rules": {
            "currency_format": True,
            "look_near": ["employers liability", "coverage b", "e.l."]
        }
    },
    "Other States Insurance - 3.C": {
        "context": ["other states insurance", "part three", "3.c", "part 3.c", "other states coverage", "part three - other states", "other states", "coverage in other states"],
        "format": "text",
        "table_headers": ["other states", "part three"],
        "special_rules": {
            "look_for_states": True
        }
    },
    "Code": {
        "context": [
            "class code", "classification code", "code no", "class codes:", "codes:",
            "classification number", "industry code", "rating classification code",
            "workers compensation class code", "ncci code", "rate code"
        ],
        "format": "code_list",
        "table_headers": ["code", "class code", "classification code", "ncci"],
        "special_rules": {
            "look_in_tables": True,
            "table_proximity": ["classification", "premium", "rate"]
        }
    },
    "Classification": {
        "context": [
            "classification description", "class description",
            "description of operations", "classification of operations",
            "class code description", "industry description",
            "business description", "operations classification",
            "work classification"
        ],
        "format": "text_list",
        "table_headers": ["classification", "description", "class description", "operations"],
        "special_rules": {
            "pair_with_codes": True,
            "exclude_common_words": True,
            "look_near_codes": True,
            "table_scan": True
        }
    },
    "Premium Basis Total Estimated Annual Remuneration": {
        "context": ["remuneration", "payroll", "premium basis", "estimated annual remuneration", "total estimated annual remuneration", "estimated total annual remuneration", "annual remuneration", "total remuneration", "estimated payroll", "total estimated payroll", "premium basis amount"],
        "format": "currency",
        "table_headers": ["premium basis", "remuneration", "payroll", "estimated annual"],
        "special_rules": {
            "sum_if_multiple": True,
            "currency_format": True,
            "look_in_tables": True
        }
    },
    "Other workplaces not shown above": {
        "context": [
            "other workplace", "additional location", "other location",
            "schedule of workplaces", "additional workplaces",
            "other insured locations", "unlisted workplaces",
            "see site location schedule", "other workplaces not shown",
            "see site schedule", "location schedule",
            "see the schedule of workplaces for this policy"
        ],
        "format": "text",
        "table_headers": ["other workplaces", "additional locations", "location"],
        "special_rules": {
            "address_format": True,
            "multi_line": True,
            "location_markers": ["location", "address", "schedule"]
        }
    },
    "Partners, Officers, Others": {
        "context": ["partners", "officers", "owners", "proprietors", "executives", "key personnel", "principals", "partners, officers and others included", "named persons"],
        "format": "text_list",
        "table_headers": ["partners", "officers", "owners", "included persons"],
        "special_rules": {
            "look_for_names": True,
            "exclude_common_words": True
        }
    }
}

# ---------------------------
# Minimal Transformation Function for CGL JSON
# ---------------------------
def transform_cgl_json_auto(cgl_json):
    """
    Converts the sharedrive CGL JSON by renaming keys to standard keys ("TAB1" and "TAB2")
    without manually writing parameter details. Assumes the JSON file has keys
    "TAB 1 - COMPARISON RESULTS" and optionally "TAB 2 - CLASSIFICATION".
    """
    if "TAB 1 - COMPARISON RESULTS" in cgl_json:
        tab1 = cgl_json["TAB 1 - COMPARISON RESULTS"]
        new_tab1 = {}
        for key, value in tab1.items():
            new_tab1[key.strip()] = value
        new_tab2 = {}
        if "TAB 2 - CLASSIFICATION" in cgl_json:
            new_tab2 = cgl_json["TAB 2 - CLASSIFICATION"]
        return {"TAB1": new_tab1, "TAB2": new_tab2}
    else:
        return cgl_json

# ---------------------------
# Transformation Function for CRPO JSON
# ---------------------------
def transform_crpo_json_auto(crpo_json):
    """
    Converts the sharedrive CRPO JSON by renaming keys to standard keys ("TAB1" and "TAB2")
    without manually writing parameter details. Assumes the JSON file has keys
    "TAB 1 - CRPO - COMPARISON RESULTS" and "TAB 2 - CPRO - COVERAGES".
    """
    if "TAB 1 - CRPO - COMPARISON RESULTS" in crpo_json:
        tab1 = crpo_json["TAB 1 - CRPO - COMPARISON RESULTS"]
        new_tab1 = {}
        for key, value in tab1.items():
            new_tab1[key.strip()] = value
        new_tab2 = {}
        if "TAB 2 - CPRO - COVERAGES" in crpo_json:
            new_tab2 = crpo_json["TAB 2 - CPRO - COVERAGES"]
        return {"TAB1": new_tab1, "TAB2": new_tab2}
    else:
        return crpo_json

# ---------------------------
# Helper Function: Load Parameters from File (for a given LOB)
# ---------------------------
def load_parameters_from_file(business_type, lob_name):
    """
    Searches for a file named <lob_name> with supported extensions in the relevant share-drive folder.
    For Commercial Line (e.g. CGL), the parameters are derived from the sharedrive file in JSON format.
    For other business types, returns the default FIELDS_TO_EXTRACT.
    """
    if business_type.upper() == "COMMERCIAL LINE":
        folder = SHAREDRIVE_COMMERCIAL
    elif business_type.upper() == "PERSONAL LINE":
        folder = SHAREDRIVE_PERSONAL
    else:
        return FIELDS_TO_EXTRACT

    for ext in ['.xlsx', '.doc', '.docx', '.txt']:
        file_path = os.path.join(folder, lob_name + ext)
        if os.path.exists(file_path):
            try:
                logging.info(f"Attempting to load parameters from {file_path}")
                if ext == '.xlsx':
                    wb = load_workbook(file_path, data_only=True)
                    ws = wb.active
                    config_json = ws["A1"].value
                    if config_json:
                        config_dict = json.loads(config_json)
                        if "TAB 1 - COMPARISON RESULTS" in config_dict:
                            return transform_cgl_json_auto(config_dict)
                        elif "TAB 1 - CRPO - COMPARISON RESULTS" in config_dict:
                            return transform_crpo_json_auto(config_dict)
                        else:
                            return config_dict
                elif ext in ('.doc', '.docx'):
                    from docx import Document
                    doc = Document(file_path)
                    full_text = "\n".join([para.text for para in doc.paragraphs])
                    config_dict = json.loads(full_text)
                    if "TAB 1 - COMPARISON RESULTS" in config_dict:
                        return transform_cgl_json_auto(config_dict)
                    elif "TAB 1 - CRPO - COMPARISON RESULTS" in config_dict:
                        return transform_crpo_json_auto(config_dict)
                    else:
                        return config_dict
                elif ext == '.txt':
                    with open(file_path, "r") as f:
                        content = f.read()
                        config_dict = json.loads(content)
                        if "TAB 1 - COMPARISON RESULTS" in config_dict:
                            return transform_cgl_json_auto(config_dict)
                        elif "TAB 1 - CRPO - COMPARISON RESULTS" in config_dict:
                            return transform_crpo_json_auto(config_dict)
                        else:
                            return config_dict
            except Exception as e:
                logging.error(f"Error loading parameters from {file_path}: {e}")
    logging.warning(f"No suitable file found or valid JSON for {lob_name}. Using default parameters.")
    return FIELDS_TO_EXTRACT

# ---------------------------
# Enhanced PDF Processing Thread
# ---------------------------
class PDFProcessingThread(QThread):
    processing_done = pyqtSignal(dict, dict, str)
    progress_update = pyqtSignal(str)

    def __init__(self, pdf1_path, pdf2_path, excel_path, field_definitions=None, parent=None):
        super().__init__(parent)
        self.pdf1_path = pdf1_path
        self.pdf2_path = pdf2_path
        self.excel_path = excel_path
        self.error_message = ""
        genai.configure(api_key='AIzaSyDtpZPgzeiuuzCBjJgCoynp_b9ufSHt-A8')
        self.model = genai.GenerativeModel('gemini-1.5-flash')
        self.field_definitions = field_definitions if field_definitions is not None else FIELDS_TO_EXTRACT
        self.enhanced_ocr = EnhancedOCR()

    def run(self):
        try:
            self.progress_update.emit("Starting PDF processing...")
            
            if isinstance(self.field_definitions, dict):
                # Distinguish whether this is a CGL route vs. CRPO route vs. generic route vs. default workers comp
                if "TAB1" in self.field_definitions and "TAB2" in self.field_definitions and "PARAMETERS" in self.field_definitions["TAB2"]:
                    # Check if this is CRPO route by examining the structure of TAB2
                    is_crpo = False
                    if self.field_definitions["TAB2"]["PARAMETERS"] and isinstance(self.field_definitions["TAB2"]["PARAMETERS"], list):
                        sample_param = self.field_definitions["TAB2"]["PARAMETERS"][0] if self.field_definitions["TAB2"]["PARAMETERS"] else {}
                        if isinstance(sample_param, dict) and "Subject of Insurance" in sample_param:
                            is_crpo = True
                    
                    if is_crpo:
                        # This is CRPO route
                        self.progress_update.emit("Processing CRPO documents...")
                        data1 = self.upload_and_query_pdf_crpo(self.pdf1_path)
                        if not data1:
                            raise Exception("Failed to process first PDF")
                        data2 = self.upload_and_query_pdf_crpo(self.pdf2_path)
                        if not data2:
                            raise Exception("Failed to process second PDF")
                        success = self.export_to_excel_crpo(data1, data2)
                        if not success:
                            raise Exception("Failed to export results to Excel for CRPO")
                    else:
                        # This is CGL route
                        self.progress_update.emit("Processing CGL documents...")
                        data1 = self.upload_and_query_pdf_cgl(self.pdf1_path)
                        if not data1:
                            raise Exception("Failed to process first PDF")
                        data2 = self.upload_and_query_pdf_cgl(self.pdf2_path)
                        if not data2:
                            raise Exception("Failed to process second PDF")
                        success = self.export_to_excel_cgl(data1, data2)
                        if not success:
                            raise Exception("Failed to export results to Excel for CGL")
                elif "excel_config" in self.field_definitions:
                    # Generic LOB route using dynamic JSON configuration
                    self.progress_update.emit("Processing generic LOB documents...")
                    data1 = self.upload_and_query_pdf(self.pdf1_path)
                    if not data1:
                        raise Exception("Failed to process first PDF")
                    data2 = self.upload_and_query_pdf(self.pdf2_path)
                    if not data2:
                        raise Exception("Failed to process second PDF")
                    success = self.export_to_excel_generic(data1, data2)
                    if not success:
                        raise Exception("Failed to export results to Excel for Generic LOB")
                else:
                    # Workers Comp route
                    self.progress_update.emit("Processing Workers Compensation documents...")
                    data1 = self.upload_and_query_pdf(self.pdf1_path)
                    if not data1:
                        raise Exception("Failed to process first PDF")
                    data2 = self.upload_and_query_pdf(self.pdf2_path)
                    if not data2:
                        raise Exception("Failed to process second PDF")
                    success = self.export_to_excel(data1, data2)
                    if not success:
                        raise Exception("Failed to export results to Excel")
            else:
                raise Exception("Invalid field definitions format")
            
            self.progress_update.emit("Processing completed successfully!")
            self.processing_done.emit(data1, data2, "")
        except Exception as e:
            self.progress_update.emit(f"Error: {str(e)}")
            self.processing_done.emit({}, {}, str(e))

    def extract_text_from_pdf(self, file_path):
        """Enhanced text extraction with OCR fallback for challenging documents"""
        try:
            self.progress_update.emit(f"Extracting text from {os.path.basename(file_path)}...")
            
            doc = fitz.open(file_path)
            pages = list(doc)
            max_workers = (os.cpu_count() or 4) * 16
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                results = list(executor.map(self.process_page_enhanced, pages))
            
            text_by_page = [res[0] for res in results]
            tables_by_page = []
            for res in results:
                tables_by_page.extend(res[1])
            
            doc.close()
            
            # Combine all text
            combined_text = "\n".join(text_by_page)
            
            # If text extraction was poor, try OCR on the entire document
            if len(combined_text.strip()) < 100:  # Very little text extracted
                self.progress_update.emit("Text extraction was limited, applying OCR...")
                ocr_text = self.apply_ocr_to_pdf(file_path)
                if len(ocr_text) > len(combined_text):
                    combined_text = ocr_text
                    logging.info("OCR provided better text extraction results")
            
            return {'text': combined_text, 'tables': tables_by_page}
            
        except Exception as e:
            logging.error(f"Error extracting text from PDF: {e}")
            # Fallback to OCR if regular extraction fails
            try:
                self.progress_update.emit("Fallback to OCR processing...")
                ocr_text = self.apply_ocr_to_pdf(file_path)
                return {'text': ocr_text, 'tables': []}
            except Exception as ocr_error:
                logging.error(f"OCR fallback also failed: {ocr_error}")
                return {'text': '', 'tables': []}

    def process_page_enhanced(self, page):
        """Enhanced page processing with OCR fallback"""
        try:
            # Try regular text extraction first
            try:
                text_dict = page.get_text("dict", sort=True)
            except Exception as e:
                logging.error(f"Page text extraction error: {e}")
                text_dict = {"blocks": []}
            
            blocks = text_dict.get("blocks", [])
            blocks.sort(key=lambda b: (b.get("bbox", [0, 0])[1], b.get("bbox", [0, 0])[0]))
            
            page_text = []
            current_section = []
            last_y = None
            last_font = None
            last_size = None

            for block in blocks:
                lines = block.get("lines", [])
                if not lines:
                    continue
                for line in lines:
                    spans = sorted(line.get("spans", []), key=lambda s: s.get("bbox", [0, 0])[0])
                    if not spans:
                        continue
                    current_y = line.get("bbox", [0, 0])[1]
                    current_font = spans[0].get("font", "")
                    current_size = spans[0].get("size", 0)
                    if last_y is not None:
                        y_gap = current_y - last_y
                        if y_gap > 15 or (current_font != last_font and y_gap > 5) or (last_size and abs(current_size - last_size) > 2):
                            if current_section:
                                page_text.append(" ".join(current_section))
                                current_section = []
                    line_text = " ".join(
                        span.get("text", "").strip()
                        for span in spans if span.get("text", "").strip()
                    )
                    if line_text:
                        current_section.append(line_text)
                    last_y = current_y
                    last_font = current_font
                    last_size = current_size
            
            if current_section:
                page_text.append(" ".join(current_section))

            extracted_text = "\n".join(page_text)
            
            # If very little text was extracted, try OCR on this page
            if len(extracted_text.strip()) < 50:
                try:
                    # Get page as image
                    mat = fitz.Matrix(2.0, 2.0)  # Higher resolution
                    pix = page.get_pixmap(matrix=mat)
                    img_data = pix.tobytes("png")
                    
                    # Convert to PIL Image
                    image = Image.open(io.BytesIO(img_data))
                    
                    # Apply OCR
                    ocr_text = self.enhanced_ocr.extract_text_with_ocr(image)
                    if len(ocr_text.strip()) > len(extracted_text.strip()):
                        extracted_text = ocr_text
                        logging.info(f"OCR improved text extraction for page")
                
                except Exception as e:
                    logging.error(f"OCR processing failed for page: {e}")

            # Extract tables
            tables_by_page = []
            try:
                found_tables = page.find_tables(
                    vertical_strategy="text",
                    horizontal_strategy="lines",
                    snap_tolerance=3,
                    join_tolerance=3,
                    edge_min_length=3,
                    min_words_vertical=2
                )
                for table in found_tables:
                    try:
                        extracted = table.extract()
                        if extracted and len(extracted) > 1:
                            cleaned_table = []
                            for row in extracted:
                                cleaned_row = [str(cell).strip() for cell in row if str(cell).strip()]
                                if cleaned_row:
                                    cleaned_table.append(cleaned_row)
                            if cleaned_table:
                                tables_by_page.append(cleaned_table)
                    except Exception as e:
                        if "not a textpage" in str(e):
                            continue
                        else:
                            logging.error(f"Table extraction error on a table: {e}")
            except Exception as e:
                if "not a textpage" in str(e):
                    pass
                else:
                    logging.error(f"Table extraction error on page: {e}")

            return (extracted_text, tables_by_page)
            
        except Exception as e:
            logging.error(f"Error processing page: {e}")
            return ("", [])

    def apply_ocr_to_pdf(self, file_path):
        """Apply OCR to entire PDF document"""
        try:
            doc = fitz.open(file_path)
            all_text = []
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Convert page to high-resolution image
                mat = fitz.Matrix(3.0, 3.0)  # High resolution for better OCR
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # Convert to PIL Image
                image = Image.open(io.BytesIO(img_data))
                
                # Enhance image quality
                enhanced_image = self.enhanced_ocr.enhance_image_quality(image)
                
                # Extract text with OCR
                page_text = self.enhanced_ocr.extract_text_with_ocr(enhanced_image)
                if page_text.strip():
                    all_text.append(page_text)
            
            doc.close()
            return "\n".join(all_text)
            
        except Exception as e:
            logging.error(f"Error applying OCR to PDF: {e}")
            return ""

    # ---------------------------
    # Workers Comp extraction
    # ---------------------------
    def upload_and_query_pdf(self, file_path):
        try:
            content = self.extract_text_from_pdf(file_path)
            if not content['text']:
                raise Exception("Failed to extract content from PDF")
            final_data = self.ai_extract_all_fields(content)
            return final_data
        except Exception as e:
            logging.error(f"Error during PDF upload/query: {e}")
            return {}

    def ai_extract_all_fields(self, content):
        table_data_summary = self.process_table_data_all(content['tables'])
        base_prompt = f"""
You are an expert insurance document analyzer for workers compensation policies.
Perform an advanced and thorough analysis of the document. The document may contain the required data anywhere within the text or tables.
Extract the following fields and return ONLY a valid JSON object with exactly these keys (even if a value is empty):

{', '.join(self.field_definitions.keys())}

Use the following synonyms for the parameters if they appear in the document:
- Policy Number: "Policy No.", "Policy Identifier", "Coverage ID"
- Renewal of Policy: "Policy Renewal", "Nonrenewal", "Renewal Effective Date", "Renewal Date", "Policy Renewal Date"
- Policy Period From: "Effective Date", "Start Date"
- Policy To: "Expiration Date", "End Date"
- Coverage Provided By: "Underwriting Company", "Insurance Provider", "Carrier"
- The Insured's Name: "Named Insured", "Policyholder", "Insured Entity"
- DBA Name: "Doing Business As (DBA)", "Trade Name"
- Mailing Address: "Insured Address", "Policyholder Address"
- Type of Entity: "Legal Entity Type", "Business Type"
- FEIN Number: "Federal Employer Identification Number", "Tax ID"
- Total Estimated Annual Premium: "Total Premium", "Annual Premium Estimate"
- Workers Compensation Insurance - 3.A: "Workers Compensation Coverage", "Coverage Type: 3.A"
- Employers Liability Insurance - 3.B: Always output exactly "WORKERS COMPENSATION AND EMPLOYERS LIABILITY INSURANCE POLICY"
- Bodily Injury by Accident - Each Accident: "Accident Coverage Limit", "Bodily Injury Per Accident"
- Bodily Injury by Disease - Policy Limit: "Disease Coverage Policy Limit", "Bodily Injury by Disease - Maximum Limit"
- Bodily Injury by Disease - Each Employee: "Per Employee Disease Coverage", "Individual Employee Injury Coverage"
- Other States Insurance - 3.C: "Other States Coverage", "Multi-State Insurance - 3.C"
  (Extract only state abbreviations or codes such as ND, OH, WA, WY; ignore any other descriptive text.)
- Code: "Classification Code", "Coverage Code", "Employer Code"
  (Extract the full classification codes from anywhere in the document, ignoring extraneous text. 
   Do not use any predetermined regex or pattern.)
- Classification: "Job Classification", "Policy Classification"
  (Extract only actual job/work classifications; 
   specifically exclude these items if they appear:
     "INCREASED LIMITS OF EMPLOYERS LIABILITY",
     "TO EQUAL MINIMUM PREMIUM (E L)",
     "PREMIUM DISCOUNT",
     "EXPENSE CONSTANT",
     "TERRORISM ACT SURCHARGE")
- Premium Basis Total Estimated Annual Remuneration: "Remuneration Basis", "Annual Payroll Estimate"
  (If multiple amounts exist, list them as comma-separated values.)
- Other workplaces not shown above: "Additional Locations", "Unlisted Workplaces"
  (Extract only actual addresses or location details; if none exist, return "N/A")
- Partners, Officers, Others: "Partners, Executives, and Others", "Exempt Individuals"

Additional instructions:
- For "Renewal of Policy", extract only the raw renewal reference (ignore any prefixes). If not found, return "N/A".
- For "DBA Name", if the document explicitly includes "DBA" followed by a value, extract that; otherwise, extract the official trade name.
- For "Employers Liability Insurance - 3.B", ignore the document content and always output exactly "WORKERS COMPENSATION AND EMPLOYERS LIABILITY INSURANCE POLICY".
- For "Other workplaces not shown above", ignore placeholders.
- For "Classification", exclude unwanted phrases as specified.
- For "Other States Insurance - 3.C", return only state abbreviations.
- For "Code", extract the full codes ignoring extraneous text.
- For "Premium Basis Total Estimated Annual Remuneration", if multiple amounts exist, list them in one string separated by commas.

Additional table context:
{table_data_summary}

Document text:
{content['text']}

Return only a valid JSON object.
"""
        final_results = {}
        try:
            response = self.model.generate_content(base_prompt)
            raw_text = response.text.strip()
            if not raw_text.startswith('{'):
                start = raw_text.find('{')
                end = raw_text.rfind('}') + 1
                if start != -1 and end != -1:
                    raw_text = raw_text[start:end]
                else:
                    raise Exception("No valid JSON found in response")
            json_data = json.loads(raw_text)
            for field_name in self.field_definitions.keys():
                raw_val = json_data.get(field_name, "N/A") or "N/A"
                validated_val = self.post_validation(raw_val, field_name)
                final_results[field_name] = validated_val
        except Exception as e:
            logging.error(f"AI extraction error: {e}")
            final_results = {fn: "N/A" for fn in self.field_definitions.keys()}
        return final_results

    def post_validation(self, value, field_name):
        if field_name == "Employers Liability Insurance - 3.B":
            return "WORKERS COMPENSATION AND EMPLOYERS LIABILITY INSURANCE POLICY"
        if not value or (isinstance(value, str) and value.strip().lower() == "n/a"):
            return "N/A"
        field_info = self.field_definitions.get(field_name, {})
        field_format = field_info.get('format', '')
        if field_format == 'date':
            return self.validate_date(value)
        elif field_format == 'currency':
            return self.validate_currency(value)
        else:
            return self.clean_text(value)

    def validate_date(self, val):
        val = val.strip()
        date_formats = ['%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%m/%d/%y', '%b %d %Y', '%B %d %Y', '%d/%m/%Y']
        for fmt in date_formats:
            try:
                parsed = datetime.strptime(val, fmt)
                return parsed.strftime('%m/%d/%Y')
            except Exception:
                continue
        return val

    def validate_currency(self, val):
        val = val.strip()
        if not val.startswith('$'):
            val = '$' + val
        try:
            float_val = float(val.replace('$', '').replace(',', ''))
            return f"${float_val:,.2f}"
        except Exception:
            return val

    def clean_text(self, val):
        if not isinstance(val, str):
            val = str(val)
        return ' '.join(val.split())

    def process_table_data_all(self, tables):
        if not tables:
            return ""
        summary_lines = []
        for i, table in enumerate(tables, start=1):
            table_text = []
            for row in table:
                row_text = " | ".join(str(cell).strip() for cell in row if str(cell).strip())
                if row_text:
                    table_text.append(row_text)
            if table_text:
                summary_lines.extend([f"\nTable {i}:"] + table_text)
        return "\n".join(summary_lines)

    # ---------------------------
    # CGL extraction
    # ---------------------------
    def upload_and_query_pdf_cgl(self, file_path):
        try:
            content = self.extract_text_from_pdf(file_path)
            if not content['text']:
                raise Exception("Failed to extract content from PDF")
            final_data = self.ai_extract_cgl_fields(content)
            return final_data
        except Exception as e:
            logging.error(f"Error during CGL PDF upload/query: {e}")
            return {}

    def ai_extract_cgl_fields(self, content):
        """
        Enhanced to incorporate:
          - Professional Liability - Limit 1
          - Professional Liability - Limit 2
          - Sexual/Physical Abuse or Molestation - Limit 1
          - Sexual/Physical Abuse or Molestation - Limit 2
          - Additional Insured – Blanket or Specific (include name + address)
          - Waiver of Subrogation – Blanket or Specific (from the ADITIONAL/OTHER INTEREST section)
          - Policy forms / endorsements with full writeup and numbering

        IMPORTANT: Also ensure that **all classification lines** (if multiple) are fully captured and listed 
        under TAB2 -> PARAMETERS so that none are missing. 
        """
        table_data_summary = self.process_table_data_all(content['tables'])
        reference_json = json.dumps(self.field_definitions, indent=2)

        prompt = f"""
You are an expert insurance document analyzer for Commercial General Liability (CGL) policies.
The data must be extracted and organized into two tabs according to the exact JSON format below.
NOTE: The extraction must include new parameters as specified and enhanced detail for certain fields.
For TAB1, please ensure the following modifications:
- Include all parameters. Under the "ADDITIONAL/OTHER INTEREST" section, add the parameter "Waiver of Subrogation – Blanket or Specific" if present.
- Enhance data extraction for "Professional Liability - Limit 1" and "Professional Liability - Limit 2" with more detailed analysis and ensure you capture all relevant numeric and descriptive details.
- Enhance data extraction for "Sexual/Physical Abuse or Molestation - Limit 1" and "Sexual/Physical Abuse or Molestation - Limit 2", capturing detailed limits and descriptive text.
- For "POLICY FORMS / CONDITIONS", in addition to listing each policy form or endorsement on its own line, also fetch any accompanying writeup along with the number (for example, "MJIL 1000 (06-10) Signature Page" or "MD 001 (07-02) Commercial Lines Policy Declarations").

For TAB2, under the "PARAMETERS" array for classification data, ensure that for each location the full classification writeup is fetched completely without truncation. For example, instead of:
"Loc#    Classification
001-001 Schools-Faculty Liability for Corporal Punishment of Students"
it should be:
"Loc#    Classification
001-001 Schools-Faculty Liability for Corporal Punishment of Students Products-completed operations are subject to the General Aggregate Limit"

IMPORTANT: do not skip or lose any classification lines. If multiple lines appear, list them all. No data should go missing.

JSON Format (sample reference):
{reference_json}

Instructions / synonyms for specific fields in TAB1 -> PARAMETERS:
1) "Professional Liability - Limit 1": synonyms may include "E&O limit 1", "Miscellaneous Professional Liability limit 1", etc.
2) "Professional Liability - Limit 2": synonyms may include "E&O limit 2", "Miscellaneous Professional Liability limit 2", etc.
3) "Additional Insured – Blanket or Specific": synonyms may include "Additional Insured Blanket" or "AI Blanket or Specific". If the document lists a name & address under Additional Insured, capture both distinctly.
4) "Waiver of Subrogation – Blanket or Specific": extract this parameter if present under "ADDITIONAL/OTHER INTEREST".
5) "Sexual/Physical Abuse or Molestation - Limit 1": capture detailed limits and descriptive text.
6) "Sexual/Physical Abuse or Molestation - Limit 2": capture detailed limits and descriptive text.
7) For TAB2 -> PARAMETERS, for the parameters "Rate -Prem/Ops" and "Premium -Prem/Ops", consider the synonym "all other".
8) For TAB2 -> PARAMETERS, for the parameters "Rate -Products" and "Premium -Products", consider the synonym "Pr/Co".

Perform a thorough analysis of the PDF text and tables below:
{table_data_summary}

Document text:
{content['text']}

Return only a valid JSON object in the format:
{{
  "TAB1": {{
    "PARAMETERS": {{ ... }},
    "LOCATIONS": [],
    "LIMITS": {{ }},
    "DEDUCTIBLE": {{ }},
    "ADDITIONAL/OTHER INTEREST": {{ }},
    "POLICY FORMS / CONDITIONS": {{ }},
    "ADDITIONAL COMMENTS": ""
  }},
  "TAB2": {{
    "PARAMETERS": []
  }}
}}
Ensure that each "POLICY FORMS / CONDITIONS" item is on its own line or array entry, and that classification writeups in TAB2 are fully fetched without truncation or omission.
"""
        final_results = {}
        try:
            response = self.model.generate_content(prompt)
            raw_text = response.text.strip()
            if not raw_text.startswith('{'):
                start = raw_text.find('{')
                end = raw_text.rfind('}') + 1
                if start != -1 and end != -1:
                    raw_text = raw_text[start:end]
                else:
                    raise Exception("No valid JSON found in response")
            final_results = json.loads(raw_text)
        except Exception as e:
            logging.error(f"CGL AI extraction error: {e}")
            # Fallback: return an empty structure matching the reference style
            tab1 = self.field_definitions.get("TAB1", {})
            parameters = {}
            if isinstance(tab1, dict):
                parameters = tab1.get("PARAMETERS", {})
            final_results = {
                "TAB1": {
                    "PARAMETERS": {key: "" for key in parameters.keys()},
                    "LOCATIONS": [],
                    "LIMITS": {},
                    "DEDUCTIBLE": {},
                    "ADDITIONAL/OTHER INTEREST": {},
                    "POLICY_FORMS / CONDITIONS": {},
                    "ADDITIONAL COMMENTS": ""
                },
                "TAB2": {"PARAMETERS": []}
            }
        return final_results

    # ---------------------------
    # CRPO extraction
    # ---------------------------
    def upload_and_query_pdf_crpo(self, file_path):
        try:
            content = self.extract_text_from_pdf(file_path)
            if not content['text']:
                raise Exception("Failed to extract content from PDF")
            final_data = self.ai_extract_crpo_fields(content)
            return final_data
        except Exception as e:
            logging.error(f"Error during CRPO PDF upload/query: {e}")
            return {}

    def ai_extract_crpo_fields(self, content):
        """
        Extract fields for Commercial Property (CRPO) policies.
        The data is organized into two tabs:
        - TAB1: General policy information, locations, additional interests, and policy forms
        - TAB2: Coverage details including location-specific and blanket coverages
        """
        table_data_summary = self.process_table_data_all(content['tables'])
        reference_json = json.dumps(self.field_definitions, indent=2)

        prompt = f"""
You are an expert insurance document analyzer for Commercial Property (CRPO) policies.
The data must be extracted and organized into two tabs according to the exact JSON format below.

For TAB1, extract the following sections:
- PARAMETERS: Basic policy information (Policy Number, Renewal of Policy, Policy Period From/To, etc.)
- LOCATIONS: List of all property locations mentioned in the document
- ADDITIONAL/OTHER INTEREST: Information about Loss Payee/Mortgagee
- POLICY FORMS / CONDITIONS: Schedule of forms and endorsements
- ADDITIONAL COMMENTS: Any other relevant information

For TAB2, extract coverage details in the "PARAMETERS" array, with each item containing:
- Loc#: Location number
- Bldg: Building number or identifier
- Subject of Insurance: Description of the insured property
- Limit: Coverage limit amount
- Co-Ins %: Co-insurance percentage
- Valuation: Valuation method (Replacement Cost, Actual Cash Value, etc.)
- Cause of Loss: Type of loss covered (Basic, Broad, Special, etc.)
- AOP Ded.: All Other Perils deductible
- W/H Ded.: Wind/Hail deductible

Organize the TAB2 data into two sections if present:
1. Location-specific coverages
2. Blanket coverages

JSON Format (sample reference):
{reference_json}

Instructions for specific fields:
1) "Policy Number": Look for policy number, policy ID, or similar identifiers
2) "Renewal of Policy": Look for renewal information, prior policy references
3) "Policy Period From/To": Look for effective dates, policy term dates
4) "Coverage Provided By": Look for insurance company, carrier, insurer
5) "The Insured's Name": Look for named insured, policyholder
6) "DBA Name": Look for "doing business as", trade name
7) "Mailing Address": Look for insured's address, mailing address
8) "Type of Entity": Look for business type, organization type
9) "Full Term Premium": Look for total premium, annual premium

For TAB2 coverages, pay special attention to tables that list property coverages.
Distinguish between location-specific coverages and blanket coverages.

Perform a thorough analysis of the PDF text and tables below:
{table_data_summary}

Document text:
{content['text']}

Return only a valid JSON object in the format:
{{
  "TAB1": {{
    "PARAMETERS": {{ ... }},
    "LOCATIONS": [],
    "ADDITIONAL/OTHER INTEREST": {{ }},
    "POLICY FORMS / CONDITIONS": {{ }},
    "ADDITIONAL COMMENTS": ""
  }},
  "TAB2": {{
    "PARAMETERS": []
  }}
}}
Ensure that all coverage details are fully captured without omission.
"""
        final_results = {}
        try:
            response = self.model.generate_content(prompt)
            raw_text = response.text.strip()
            if not raw_text.startswith('{'):
                start = raw_text.find('{')
                end = raw_text.rfind('}') + 1
                if start != -1 and end != -1:
                    raw_text = raw_text[start:end]
                else:
                    raise Exception("No valid JSON found in response")
            final_results = json.loads(raw_text)
        except Exception as e:
            logging.error(f"CRPO AI extraction error: {e}")
            # Fallback: return an empty structure matching the reference style
            tab1 = self.field_definitions.get("TAB1", {})
            parameters = {}
            if isinstance(tab1, dict):
                parameters = tab1.get("PARAMETERS", {})
            final_results = {
                "TAB1": {
                    "PARAMETERS": {key: "" for key in parameters.keys()},
                    "LOCATIONS": [],
                    "ADDITIONAL/OTHER INTEREST": {},
                    "POLICY FORMS / CONDITIONS": {},
                    "ADDITIONAL COMMENTS": ""
                },
                "TAB2": {"PARAMETERS": []}
            }
        return final_results

    # ---------------------------
    # Excel Export for Workers Compensation
    # ---------------------------
    def export_to_excel(self, data1, data2):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Comparison Results"

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            subheader_font = Font(name='Calibri', size=11, bold=True, color='000000')
            header_fill = PatternFill(start_color="002175", end_color="002175", fill_type='solid')
            subheader_fill = PatternFill(start_color="FF8A00", end_color="FF8A00", fill_type='solid')
            alt_row_fill = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = ["Sl. No.", "SECTION", "Prior Term", "Current Term", "Status"]
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            sheet.cell(row=2, column=3, value="Dec Page")
            sheet.cell(row=2, column=4, value="Renewal")
            for col in [3, 4]:
                cell = sheet.cell(row=2, column=col)
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            row_index = 3
            for idx, field_name in enumerate(self.field_definitions.keys(), start=1):
                row_fill = alt_row_fill if idx % 2 == 0 else None

                cell = sheet.cell(row=row_index, column=1, value=idx)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                cell = sheet.cell(row=row_index, column=2, value=field_name)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                prior_val = data1.get(field_name, "N/A")
                cell = sheet.cell(row=row_index, column=3, value=prior_val)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                current_val = data2.get(field_name, "N/A")
                cell = sheet.cell(row=row_index, column=4, value=current_val)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                status = "MATCHES" if prior_val == current_val else "DOESN'T MATCH"
                cell = sheet.cell(row=row_index, column=5, value=status)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

                row_index += 1

            column_widths = {
                'A': 8,
                'B': 35,
                'C': 45,
                'D': 45,
                'E': 15
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            sheet.row_dimensions[1].height = 30
            sheet.row_dimensions[2].height = 25
            for i in range(3, sheet.max_row + 1):
                sheet.row_dimensions[i].height = 35

            sheet.merge_cells('A1:A2')
            sheet.merge_cells('B1:B2')
            sheet.merge_cells('E1:E2')

            workbook.save(self.excel_path)
            return True
        except Exception as e:
            logging.error(f"Error in export_to_excel: {str(e)}")
            raise Exception(f"Error exporting to Excel: {str(e)}")

    # ---------------------------
    # Utility to stringify complex data (dict/list) for Excel
    # ---------------------------
    def _stringify(self, value):
        if isinstance(value, dict):
            return "; ".join(f"{k}: {v}" for k, v in value.items())
        elif isinstance(value, list):
            return "\n".join(self._stringify(x) for x in value)
        else:
            return str(value)

    # ---------------------------
    # Excel Export for CGL
    # ---------------------------
    def export_to_excel_cgl(self, data1, data2):
        """
        Creates two sheets:
          1) "CGL - Comparison Results" (TAB1)
          2) "CGL - Classification" (TAB2)
        We keep existing structure, ensuring we handle newly added fields (like professional liability, 
        sexual abuse limits, additional insured, etc.) but do NOT alter how Tab1 is displayed, 
        and ensure no classification is lost from Tab2.
        """
        try:
            # Retrieve and validate data for TAB1 and TAB2
            t1_prior = data1.get("TAB1", {})
            t1_current = data2.get("TAB1", {})
            if not isinstance(t1_prior, dict):
                t1_prior = {}
            if not isinstance(t1_current, dict):
                t1_current = {}

            tab2_prior = data1.get("TAB2", {})
            tab2_current = data2.get("TAB2", {})
            if not isinstance(tab2_prior, dict):
                tab2_prior = {}
            if not isinstance(tab2_current, dict):
                tab2_current = {}

            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            blue_fill = PatternFill(start_color="002175", end_color="002175", fill_type='solid')
            orange_fill = PatternFill(start_color="FF8A00", end_color="FF8A00", fill_type='solid')
            alt_row_fill = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # --- Export TAB1 - Comparison Results ---
            tab1_sheet = workbook.create_sheet("CGL - Comparison Results")
            # Header row for tab1
            tab1_headers = ["Sl. No.", "Section", "Parameter", "Prior Term", "Current Term", "Status"]
            for col, h in enumerate(tab1_headers, start=1):
                cell = tab1_sheet.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            row_idx = 2
            sl_no = 1
            # Iterate over each section in t1_prior
            for section, content in t1_prior.items():
                # Section header row
                tab1_sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
                sec_cell = tab1_sheet.cell(row=row_idx, column=2, value=section)
                sec_cell.font = header_font
                sec_cell.fill = orange_fill
                sec_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sec_cell.border = thin_border
                tab1_sheet.cell(row=row_idx, column=1, value="").border = thin_border
                row_idx += 1

                if isinstance(content, dict):
                    for param, prior_val in content.items():
                        current_val = ""
                        if isinstance(t1_current.get(section, {}), dict):
                            current_val = t1_current.get(section, {}).get(param, "")
                        status = "MATCHES" if prior_val == current_val else "DOESN'T MATCH"
                        tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=3, value=param).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=4, value=str(prior_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=5, value=str(current_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                        row_idx += 1
                        sl_no += 1
                elif isinstance(content, list):
                    for i, item in enumerate(content, start=1):
                        current_list = t1_current.get(section, [])
                        current_val = current_list[i - 1] if i - 1 < len(current_list) else ""
                        status = "MATCHES" if item == current_val else "DOESN'T MATCH"
                        tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=3, value=f"Item {i}").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=4, value=self._stringify(item)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=5, value=self._stringify(current_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                        row_idx += 1
                        sl_no += 1
                else:
                    current_val = t1_current.get(section, "")
                    status = "MATCHES" if content == current_val else "DOESN'T MATCH"
                    tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                    tab1_sheet.cell(row=row_idx, column=3, value="Value").border = thin_border
                    tab1_sheet.cell(row=row_idx, column=4, value=str(content)).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=5, value=str(current_val)).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                    row_idx += 1
                    sl_no += 1

            tab1_sheet.column_dimensions['A'].width = 8
            tab1_sheet.column_dimensions['B'].width = 20
            tab1_sheet.column_dimensions['C'].width = 30
            tab1_sheet.column_dimensions['D'].width = 40
            tab1_sheet.column_dimensions['E'].width = 40
            tab1_sheet.column_dimensions['F'].width = 15

            # --- Export TAB2 - Classification ---
            tab2_sheet = workbook.create_sheet("CGL - Classification")
            tab2_sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=10)
            tab2_sheet.merge_cells(start_row=1, start_column=11, end_row=1, end_column=19)

            cell = tab2_sheet.cell(row=1, column=1, value="Sl. No.")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell = tab2_sheet.cell(row=1, column=2, value="Prior Term")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell = tab2_sheet.cell(row=1, column=11, value="Current Term")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell = tab2_sheet.cell(row=1, column=20, value="Status")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

            sub_headers = ["Loc#", "Classification", "Code", "Exposure", "Basis", "Rate -Prem/Ops", "Rate -Products", "Premium - Prem/Ops", "Premium - Products"]
            for j, sh in enumerate(sub_headers, start=2):
                cell = tab2_sheet.cell(row=2, column=j, value=sh)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            for j, sh in enumerate(sub_headers, start=11):
                cell = tab2_sheet.cell(row=2, column=j, value=sh)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            prior_class_list = tab2_prior.get("PARAMETERS", [])
            current_class_list = tab2_current.get("PARAMETERS", [])
            if not isinstance(prior_class_list, list):
                prior_class_list = []
            if not isinstance(current_class_list, list):
                current_class_list = []

            max_len_class = max(len(prior_class_list), len(current_class_list))
            row_idx_class = 3
            for i in range(max_len_class):
                tab2_sheet.cell(row=row_idx_class, column=1, value=(i + 1)).border = thin_border
                p_rec = prior_class_list[i] if i < len(prior_class_list) else {}
                c_rec = current_class_list[i] if i < len(current_class_list) else {}
                for j, key in enumerate(sub_headers, start=2):
                    val = p_rec.get(key, "")
                    cell = tab2_sheet.cell(row=row_idx_class, column=j, value=self._stringify(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for j, key in enumerate(sub_headers, start=11):
                    val = c_rec.get(key, "")
                    cell = tab2_sheet.cell(row=row_idx_class, column=j, value=self._stringify(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                status = "MATCHES" if p_rec == c_rec else "DOESN'T MATCH"
                cell = tab2_sheet.cell(row=row_idx_class, column=20, value=status)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                row_idx_class += 1

            col_widths = {
                1: 8, 2: 12, 3: 25, 4: 10, 5: 12, 6: 12, 7: 12, 8: 15, 9: 15, 10: 12,
                11: 12, 12: 25, 13: 10, 14: 12, 15: 12, 16: 12, 17: 15, 18: 15, 19: 12, 20: 15
            }
            for col_i, width in col_widths.items():
                tab2_sheet.column_dimensions[chr(64 + col_i)].width = width

            workbook.save(self.excel_path)
            return True
        except Exception as e:
            logging.error(f"Error exporting to Excel for CGL: {str(e)}")
            raise Exception(f"Error exporting to Excel for CGL: {str(e)}")

    # ---------------------------
    # Excel Export for CRPO
    # ---------------------------
    def export_to_excel_crpo(self, data1, data2):
        """
        Creates two sheets:
          1) "CRPO - Comparison Results" (TAB1)
          2) "CRPO - Coverages" (TAB2)
        
        The first sheet shows general policy information, locations, and other details.
        The second sheet shows property coverages with location-specific and blanket coverages.
        """
        try:
            # Retrieve and validate data for TAB1 and TAB2
            t1_prior = data1.get("TAB1", {})
            t1_current = data2.get("TAB1", {})
            if not isinstance(t1_prior, dict):
                t1_prior = {}
            if not isinstance(t1_current, dict):
                t1_current = {}

            tab2_prior = data1.get("TAB2", {})
            tab2_current = data2.get("TAB2", {})
            if not isinstance(tab2_prior, dict):
                tab2_prior = {}
            if not isinstance(tab2_current, dict):
                tab2_current = {}

            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            blue_fill = PatternFill(start_color="002175", end_color="002175", fill_type='solid')
            orange_fill = PatternFill(start_color="FF8A00", end_color="FF8A00", fill_type='solid')
            alt_row_fill = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # --- Export TAB1 - Comparison Results ---
            tab1_sheet = workbook.create_sheet("CRPO - Comparison Results")
            # Header row for tab1
            tab1_headers = ["Sl. No.", "Section", "Parameter", "Prior Term", "Current Term", "Status"]
            for col, h in enumerate(tab1_headers, start=1):
                cell = tab1_sheet.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            row_idx = 2
            sl_no = 1
            # Iterate over each section in t1_prior
            for section, content in t1_prior.items():
                # Section header row
                tab1_sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
                sec_cell = tab1_sheet.cell(row=row_idx, column=2, value=section)
                sec_cell.font = header_font
                sec_cell.fill = orange_fill
                sec_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sec_cell.border = thin_border
                tab1_sheet.cell(row=row_idx, column=1, value="").border = thin_border
                row_idx += 1

                if isinstance(content, dict):
                    for param, prior_val in content.items():
                        current_val = ""
                        if isinstance(t1_current.get(section, {}), dict):
                            current_val = t1_current.get(section, {}).get(param, "")
                        status = "MATCHES" if prior_val == current_val else "DOESN'T MATCH"
                        tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=3, value=param).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=4, value=str(prior_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=5, value=str(current_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                        row_idx += 1
                        sl_no += 1
                elif isinstance(content, list):
                    for i, item in enumerate(content, start=1):
                        current_list = t1_current.get(section, [])
                        current_val = current_list[i - 1] if i - 1 < len(current_list) else ""
                        status = "MATCHES" if item == current_val else "DOESN'T MATCH"
                        tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=3, value=f"Item {i}").border = thin_border
                        tab1_sheet.cell(row=row_idx, column=4, value=self._stringify(item)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=5, value=self._stringify(current_val)).border = thin_border
                        tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                        row_idx += 1
                        sl_no += 1
                else:
                    current_val = t1_current.get(section, "")
                    status = "MATCHES" if content == current_val else "DOESN'T MATCH"
                    tab1_sheet.cell(row=row_idx, column=1, value=sl_no).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=2, value="").border = thin_border
                    tab1_sheet.cell(row=row_idx, column=3, value="Value").border = thin_border
                    tab1_sheet.cell(row=row_idx, column=4, value=str(content)).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=5, value=str(current_val)).border = thin_border
                    tab1_sheet.cell(row=row_idx, column=6, value=status).border = thin_border
                    row_idx += 1
                    sl_no += 1

            tab1_sheet.column_dimensions['A'].width = 8
            tab1_sheet.column_dimensions['B'].width = 20
            tab1_sheet.column_dimensions['C'].width = 30
            tab1_sheet.column_dimensions['D'].width = 40
            tab1_sheet.column_dimensions['E'].width = 40
            tab1_sheet.column_dimensions['F'].width = 15

            # --- Export TAB2 - Coverages ---
            tab2_sheet = workbook.create_sheet("CRPO - Coverages")
            
            # Create headers for location specific coverages
            tab2_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
            cell = tab2_sheet.cell(row=1, column=1, value="Prior Term")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            tab2_sheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=18)
            cell = tab2_sheet.cell(row=1, column=10, value="Current Term")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            cell = tab2_sheet.cell(row=1, column=19, value="Status")
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
            # Create section header for location specific coverages
            tab2_sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
            cell = tab2_sheet.cell(row=2, column=1, value="COMMERCIAL PROPERTY LOCATION SPECIFIC COVERAGES")
            cell.font = header_font
            cell.fill = orange_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            tab2_sheet.cell(row=2, column=19, value="").border = thin_border
            
            # Create column headers for location specific coverages
            coverage_headers = ["Loc.#", "Bldg.", "Subject of Insurance", "Limit", "Co-Ins %", "Valuation", "Cause of Loss", "AOP Ded.", "W/H Ded."]
            for j, header in enumerate(coverage_headers, start=1):
                cell = tab2_sheet.cell(row=3, column=j, value=header)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            for j, header in enumerate(coverage_headers, start=10):
                cell = tab2_sheet.cell(row=3, column=j, value=header)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            cell = tab2_sheet.cell(row=3, column=19, value="")
            cell.border = thin_border
            
            # Process location specific coverages
            prior_coverages = tab2_prior.get("PARAMETERS", [])
            current_coverages = tab2_current.get("PARAMETERS", [])
            
            if not isinstance(prior_coverages, list):
                prior_coverages = []
            if not isinstance(current_coverages, list):
                current_coverages = []
            
            # Fill in location specific coverages data
            row_idx = 4
            max_coverages = max(len(prior_coverages), len(current_coverages))
            
            for i in range(max_coverages):
                p_cov = prior_coverages[i] if i < len(prior_coverages) else {}
                c_cov = current_coverages[i] if i < len(current_coverages) else {}
                
                # Prior term data
                for j, key in enumerate(coverage_headers, start=1):
                    val = p_cov.get(key.replace(".", "#"), "") if key in ["Loc.#", "Bldg."] else p_cov.get(key, "")
                    cell = tab2_sheet.cell(row=row_idx, column=j, value=self._stringify(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Current term data
                for j, key in enumerate(coverage_headers, start=10):
                    val = c_cov.get(key.replace(".", "#"), "") if key in ["Loc.#", "Bldg."] else c_cov.get(key, "")
                    cell = tab2_sheet.cell(row=row_idx, column=j, value=self._stringify(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Status
                status = "MATCHES" if p_cov == c_cov else "DOESN'T MATCH"
                cell = tab2_sheet.cell(row=row_idx, column=19, value=status)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                row_idx += 1
            
            # Add a few blank rows
            for _ in range(3):
                for j in range(1, 20):
                    tab2_sheet.cell(row=row_idx, column=j, value="").border = thin_border
                row_idx += 1
            
            # Create section header for blanket coverages
            tab2_sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=18)
            cell = tab2_sheet.cell(row=row_idx, column=1, value="COMMERCIAL PROPERTY BLANKET COVERAGES")
            cell.font = header_font
            cell.fill = orange_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            tab2_sheet.cell(row=row_idx, column=19, value="").border = thin_border
            
            row_idx += 1
            
            # Create column headers for blanket coverages (same as location specific)
            for j, header in enumerate(coverage_headers, start=1):
                cell = tab2_sheet.cell(row=row_idx, column=j, value=header)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            for j, header in enumerate(coverage_headers, start=10):
                cell = tab2_sheet.cell(row=row_idx, column=j, value=header)
                cell.font = header_font
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            cell = tab2_sheet.cell(row=row_idx, column=19, value="")
            cell.border = thin_border
            
            row_idx += 1
            
            # Add a few blank rows for blanket coverages
            for _ in range(5):
                for j in range(1, 20):
                    tab2_sheet.cell(row=row_idx, column=j, value="").border = thin_border
                row_idx += 1
            
            # Set column widths
            col_widths = {
                1: 8, 2: 8, 3: 25, 4: 12, 5: 10, 6: 15, 7: 15, 8: 10, 9: 10,
                10: 8, 11: 8, 12: 25, 13: 12, 14: 10, 15: 15, 16: 15, 17: 10, 18: 10, 19: 15
            }
            
            for col_i, width in col_widths.items():
                col_letter = chr(64 + col_i) if col_i <= 26 else chr(64 + col_i // 26) + chr(64 + col_i % 26)
                tab2_sheet.column_dimensions[col_letter].width = width
            
            workbook.save(self.excel_path)
            return True
        except Exception as e:
            logging.error(f"Error exporting to Excel for CRPO: {str(e)}")
            raise Exception(f"Error exporting to Excel for CRPO: {str(e)}")

    # ---------------------------
    # Excel Export for Generic LOB
    # ---------------------------
    def export_to_excel_generic(self, data1, data2):
        try:
            excel_config = self.field_definitions.get("excel_config", {})
            sheets_config = excel_config.get("sheets", [])
            workbook = Workbook()
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if not sheets_config:
                sheet = workbook.active
                sheet.title = "Comparison Results"
                headers = ["Parameter", "Prior Term", "Current Term", "Status"]
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.border = thin_border
                row_index = 2
                all_params = sorted(set(list(data1.keys()) + list(data2.keys())))
                for param in all_params:
                    prior_val = data1.get(param, "N/A")
                    current_val = data2.get(param, "N/A")
                    status = "MATCHES" if prior_val == current_val else "DOESN'T MATCH"
                    sheet.cell(row=row_index, column=1, value=param).border = thin_border
                    sheet.cell(row=row_index, column=2, value=str(prior_val)).border = thin_border
                    sheet.cell(row=row_index, column=3, value=str(current_val)).border = thin_border
                    sheet.cell(row=row_index, column=4, value=status).border = thin_border
                    row_index += 1
            else:
                for sheet_conf in sheets_config:
                    sheet_name = sheet_conf.get("name", "Sheet")
                    sheet = workbook.create_sheet(title=sheet_name)
                    headers = sheet_conf.get("headers", [])
                    for col, header_conf in enumerate(headers, start=1):
                        header_label = header_conf.get("label", "")
                        cell = sheet.cell(row=1, column=col, value=header_label)
                        width = header_conf.get("width", None)
                        if width:
                            col_letter = chr(64 + col)
                            sheet.column_dimensions[col_letter].width = width
                        font_color = header_conf.get("font_color", "000000")
                        bg_color = header_conf.get("bg_color", "FFFFFF")
                        cell.font = Font(bold=True, color=font_color)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    section_type = sheet_conf.get("sections", "auto")
                    row_index = 2
                    if section_type == "auto":
                        all_params = sorted(set(list(data1.keys()) + list(data2.keys())))
                        for param in all_params:
                            prior_val = data1.get(param, "N/A")
                            current_val = data2.get(param, "N/A")
                            status = "MATCHES" if prior_val == current_val else "DOESN'T MATCH"
                            sheet.cell(row=row_index, column=1, value=param).border = thin_border
                            sheet.cell(row=row_index, column=2, value=str(prior_val)).border = thin_border
                            sheet.cell(row=row_index, column=3, value=str(current_val)).border = thin_border
                            sheet.cell(row=row_index, column=4, value=status).border = thin_border
                            row_index += 1
                    else:
                        pass
            workbook.save(self.excel_path)
            return True
        except Exception as e:
            logging.error(f"Error exporting to Excel for Generic LOB: {str(e)}")
            raise Exception(f"Error exporting to Excel for Generic LOB: {str(e)}")

# ---------------------------
# CustomFrame
# ---------------------------
class CustomFrame(QFrame):
    def __init__(self):
        super().__init__()
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {COLORS['background']};
                border: 2px solid {COLORS['border']};
                border-radius: 10px;
                padding: 20px;
            }}
        """)

# ---------------------------
# Excel Viewer Window
# ---------------------------
class ExcelViewerWindow(QMainWindow):
    def __init__(self, excel_file_path):
        super().__init__()
        self.excel_file_path = excel_file_path
        self.setWindowTitle("Exported Excel - Spreadsheet View")
        self.resize(800, 600)
        self.init_ui()

    def init_ui(self):
        self.table = QTableWidget()
        self.table.setShowGrid(False)
        self.table.setFrameStyle(QFrame.NoFrame)
        self.table.setStyleSheet("""
            QTableWidget {
                border: none;
            }
            QTableWidget::item {
                border: none;
            }
            QHeaderView::section {
                border: none;
            }
        """)
        layout = QVBoxLayout()
        layout.addWidget(self.table)
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
        self.load_excel()

    def load_excel(self):
        wb = load_workbook(self.excel_file_path)
        sheet = wb.active
        rows = list(sheet.iter_rows())
        if not rows:
            return
        self.table.setRowCount(len(rows))
        max_cols = max(len(row) for row in rows)
        self.table.setColumnCount(max_cols)
        headers = []
        for j in range(max_cols):
            if j < len(rows[0]):
                headers.append(str(rows[0][j].value) if rows[0][j].value is not None else "")
            else:
                headers.append("")
        self.table.setHorizontalHeaderLabels(headers)
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                value = cell.value
                item = QTableWidgetItem(str(value) if value is not None else "")
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and len(rgb) == 8:
                        color = f"#{rgb[-6:]}"
                        item.setBackground(QColor(color))
                if cell.font:
                    font = QFont()
                    if cell.font.bold:
                        font.setBold(True)
                    if cell.font.size:
                        try:
                            font_size = int(cell.font.size)
                            font.setPointSize(font_size)
                        except:
                            font.setPointSize(11)
                    item.setFont(font)
                self.table.setItem(i, j, item)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

# ---------------------------
# Welcome Page
# ---------------------------
class WelcomePage(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Marble Box - Welcome")
        self.setGeometry(100, 100, 800, 600)
        if os.path.exists("temp/icon.png"):
            self.setWindowIcon(QIcon("temp/icon.png"))
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.selected_business = None
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        top_bar = QHBoxLayout()
        back_button = QPushButton("←")
        back_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['login_secondary']};
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['login_accent']};
            }}
        """)
        back_button.clicked.connect(self.go_back)
        top_bar.addWidget(back_button, alignment=Qt.AlignLeft)
        top_bar.addStretch()

        logout_button = QPushButton("⏻")
        logout_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['login_secondary']};
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['login_accent']};
            }}
        """)
        logout_button.clicked.connect(self.logout)
        top_bar.addWidget(logout_button, alignment=Qt.AlignRight)
        main_layout.addLayout(top_bar)

        welcome_label = QLabel("WELCOME!! \n PDF COMPARISON TOOL")
        welcome_label.setAlignment(Qt.AlignCenter)
        welcome_label.setStyleSheet(f"color: {COLORS['login_primary']}; font-size: 24px; font-weight: bold;")
        main_layout.addWidget(welcome_label)
        main_layout.addSpacing(20)

        if os.path.exists("temp/logo.png"):
            pixmap = QPixmap("temp/logo.png")
        else:
            response = requests.get(LOGO_URL)
            with open("temp/logo.png", "wb") as f:
                f.write(response.content)
            pixmap = QPixmap("temp/logo.png")

        logo_label = QLabel()
        logo_label.setPixmap(pixmap.scaled(300, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(logo_label)
        main_layout.addSpacing(20)

        dropdown_frame = QFrame()
        dropdown_frame.setStyleSheet("background-color: transparent;")
        dropdown_layout = QVBoxLayout(dropdown_frame)

        self.business_combo = QComboBox()
        self.business_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: white;
                padding: 8px; 
                font-size: 14px; 
                border: 1px solid {COLORS['login_disabled']};
            }}
            QComboBox QAbstractItemView {{
                background-color: white;
                selection-background-color: {COLORS['accent']};
                selection-color: black;
                font-size: 14px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: {COLORS['accent']};
                font-weight: bold;
                color: black;
            }}
        """)
        self.business_combo.addItem("-- Select a Business Type --")
        self.business_combo.addItem("PERSONAL LINE")
        self.business_combo.addItem("COMMERCIAL LINE")
        self.business_combo.currentIndexChanged.connect(self.on_business_selected)
        dropdown_layout.addWidget(self.business_combo)

        self.lob_combo = QComboBox()
        self.lob_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: white;
                padding: 8px; 
                font-size: 14px; 
                border: 1px solid {COLORS['login_disabled']};
            }}
            QComboBox QAbstractItemView {{
                background-color: white;
                selection-background-color: {COLORS['accent']};
                selection-color: black;
                font-size: 14px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: {COLORS['accent']};
                font-weight: bold;
                color: black;
            }}
        """)
        self.lob_combo.setEnabled(False)
        dropdown_layout.addWidget(self.lob_combo)

        self.proceed_button = QPushButton("  Proceed")
        self.proceed_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['login_primary']};
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['login_accent']};
            }}
        """)
        if os.path.exists("temp/icon.png"):
            self.proceed_button.setIcon(QIcon("temp/icon.png"))
            self.proceed_button.setIconSize(QSize(20, 20))
        self.proceed_button.clicked.connect(self.on_proceed)
        dropdown_layout.addWidget(self.proceed_button)

        main_layout.addWidget(dropdown_frame)
        main_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Expanding))

    def on_business_selected(self, index):
        if index <= 0:
            self.lob_combo.clear()
            self.lob_combo.setEnabled(False)
            self.lob_combo.addItem("-- Select an Business Type --")
            self.selected_business = None
        else:
            self.selected_business = self.business_combo.currentText()
            self.lob_combo.clear()
            self.lob_combo.addItem("-- Select an LOB --")

            if self.selected_business.upper() == "PERSONAL LINE":
                folder = SHAREDRIVE_PERSONAL
            elif self.selected_business.upper() == "COMMERCIAL LINE":
                folder = SHAREDRIVE_COMMERCIAL
            else:
                folder = None

            if folder and os.path.exists(folder):
                lob_names_found = set()
                for file in os.listdir(folder):
                    if file.endswith((".xlsx", ".doc", ".docx", ".txt")):
                        lob_name = os.path.splitext(file)[0]
                        lob_names_found.add(lob_name)
                for ln in sorted(lob_names_found):
                    self.lob_combo.addItem(ln)

            self.lob_combo.setEnabled(True)

    def on_proceed(self):
        if self.business_combo.currentIndex() <= 0 or self.lob_combo.currentIndex() <= 0:
            QMessageBox.warning(self, "Error", "Please select both a Business Type and an LOB.")
            return
        self.open_pdf_comparison()

    def open_pdf_comparison(self):
        self.pdf_tool = PDFComparisonTool(self.business_combo.currentText(), self.lob_combo.currentText())
        self.pdf_tool.show()
        self.close()

    def go_back(self):
        from_index = __import__("__main__").login_window
        if from_index:
            from_index.show()
        else:
            self.login_window = LoginWindow()
            self.login_window.show()
        self.close()

    def logout(self):
        self.login_window = LoginWindow()
        self.login_window.show()
        self.close()

# ---------------------------
# Enhanced PDFComparisonTool
# ---------------------------
class PDFComparisonTool(QMainWindow):
    def __init__(self, business_type, lob_name):
        super().__init__()
        self.business_type = business_type
        self.lob_name = lob_name
        self.field_definitions = load_parameters_from_file(self.business_type, self.lob_name)
        self.pdf1_path = None
        self.pdf2_path = None
        self.excel_path = None
        self.download_resources()
        self.init_ui()

    def download_resources(self):
        try:
            if not os.path.exists('temp'):
                os.makedirs('temp')
            logo_response = requests.get(LOGO_URL)
            self.logo_path = os.path.join('temp', 'logo.png')
            with open(self.logo_path, 'wb') as f:
                f.write(logo_response.content)

            icon_response = requests.get(ICON_URL)
            self.icon_path = os.path.join('temp', 'icon.png')
            with open(self.icon_path, 'wb') as f:
                f.write(icon_response.content)
        except Exception as e:
            logging.error(f"Error downloading resources: {str(e)}")

    def init_ui(self):
        self.setWindowTitle('Marble Box PDF Comparison Tool - Enhanced')
        self.setGeometry(100, 100, 800, 600)
        self.setStyleSheet(f"background-color: {COLORS['background']};")

        if hasattr(self, 'icon_path'):
            self.setWindowIcon(QIcon(self.icon_path))

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        top_bar = QHBoxLayout()
        back_button = QPushButton("←")
        back_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['login_secondary']};
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['login_accent']};
            }}
        """)
        back_button.clicked.connect(self.go_back)
        top_bar.addWidget(back_button, alignment=Qt.AlignLeft)
        top_bar.addStretch()

        logout_button = QPushButton("LOG OUT  ⏻")
        logout_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['login_secondary']};
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['login_accent']};
            }}
        """)
        logout_button.clicked.connect(self.logout)
        top_bar.addWidget(logout_button, alignment=Qt.AlignRight)
        main_layout.addLayout(top_bar)

        if hasattr(self, 'logo_path'):
            logo_label = QLabel()
            pixmap = QPixmap(self.logo_path)
            scaled_pixmap = pixmap.scaled(300, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
            logo_label.setAlignment(Qt.AlignCenter)
            main_layout.addWidget(logo_label)

        writeup_label = QLabel(f"{self.lob_name} - Enhanced OCR & Email System")
        writeup_label.setAlignment(Qt.AlignCenter)
        writeup_label.setStyleSheet(f"color: {COLORS['primary']}; font-size: 14px; font-weight: bold;")
        main_layout.addWidget(writeup_label)

        frame = CustomFrame()
        frame_layout = QVBoxLayout(frame)

        button_style = f"""
            QPushButton {{
                background-color: {COLORS['button']};
                color: white;
                border: none;
                padding: 15px;
                border-radius: 5px;
                font-size: 14px;
                min-width: 200px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['button_hover']};
            }}
            QPushButton:disabled {{
                background-color: {COLORS['text']};
            }}
        """

        self.pdf1_button = QPushButton('Select First PDF File')
        self.pdf1_button.setStyleSheet(button_style)
        frame_layout.addWidget(self.pdf1_button)

        self.pdf2_button = QPushButton('Select Second PDF File')
        self.pdf2_button.setStyleSheet(button_style)
        frame_layout.addWidget(self.pdf2_button)

        self.excel_button = QPushButton('Select Excel Output Path')
        self.excel_button.setStyleSheet(button_style)
        frame_layout.addWidget(self.excel_button)

        self.compare_button = QPushButton('Run Enhanced Comparison')
        self.compare_button.setStyleSheet(button_style)
        frame_layout.addWidget(self.compare_button)

        self.view_excel_button = QPushButton('Preview Spreadsheet')
        self.view_excel_button.setStyleSheet(button_style)
        self.view_excel_button.setEnabled(False)
        frame_layout.addWidget(self.view_excel_button)
        self.view_excel_button.clicked.connect(self.open_excel_viewer)

        self.status_label = QLabel('Please select the first PDF file')
        self.status_label.setStyleSheet(f"color: {COLORS['primary']}; font-size: 14px; margin: 10px 0;")
        frame_layout.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: 2px solid {COLORS['border']};
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }}
            QProgressBar::chunk {{
                background-color: {COLORS['progress']};
                width: 20px;
            }}
        """)
        frame_layout.addWidget(self.progress_bar)

        main_layout.addWidget(frame)

        footer_label = QLabel('©(2025) Marble Box . All Rights Reserved - Enhanced Version')
        footer_label.setStyleSheet(f"color: {COLORS['text']}; font-size: 12px; margin-top: 10px;")
        footer_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(footer_label)

        self.pdf1_button.clicked.connect(lambda: self.select_pdf(1))
        self.pdf2_button.clicked.connect(lambda: self.select_pdf(2))
        self.excel_button.clicked.connect(self.select_excel)
        self.compare_button.clicked.connect(self.run_comparison)

        self.pdf2_button.setEnabled(False)
        self.excel_button.setEnabled(False)
        self.compare_button.setEnabled(False)

    def go_back(self):
        self.welcome_page = WelcomePage()
        self.welcome_page.show()
        self.close()

    def logout(self):
        self.login_window = LoginWindow()
        self.login_window.show()
        self.close()

    def select_pdf(self, pdf_num):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, f'Select PDF {pdf_num}', '', 'PDF Files (*.pdf)')
            if file_path:
                if pdf_num == 1:
                    self.pdf1_path = file_path
                    self.status_label.setText('PDF 1 loaded successfully')
                    self.pdf2_button.setEnabled(True)
                    self.progress_bar.setValue(30)
                else:
                    self.pdf2_path = file_path
                    self.status_label.setText('PDF 2 loaded successfully')
                    self.excel_button.setEnabled(True)
                    self.progress_bar.setValue(60)
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Error selecting PDF: {str(e)}')

    def select_excel(self):
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, 'Select Excel Output Location', '', 'Excel Files (*.xlsx)')
            if file_path:
                self.excel_path = file_path
                self.status_label.setText('Excel path selected')
                self.compare_button.setEnabled(True)
                self.progress_bar.setValue(70)
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Error selecting Excel path: {str(e)}')

    def run_comparison(self):
        try:
            if not all([self.pdf1_path, self.pdf2_path, self.excel_path]):
                raise ValueError("Please select all required files")
            self.status_label.setText('Processing PDFs with enhanced OCR... Please wait.')
            self.progress_bar.setValue(80)
            self.thread = PDFProcessingThread(self.pdf1_path, self.pdf2_path, self.excel_path, field_definitions=self.field_definitions)
            self.thread.processing_done.connect(self.on_processing_done)
            self.thread.progress_update.connect(self.on_progress_update)
            self.thread.start()
        except Exception as e:
            self.status_label.setText(f'Error: {str(e)}')
            QMessageBox.critical(self, 'Error', str(e))
            self.progress_bar.setValue(0)

    def on_progress_update(self, message):
        self.status_label.setText(message)

    def on_processing_done(self, data1, data2, error_message):
        if error_message:
            self.status_label.setText(f'Error: {error_message}')
            QMessageBox.critical(self, 'Error', error_message)
            self.progress_bar.setValue(0)
        else:
            self.status_label.setText('Enhanced comparison completed successfully!')
            self.progress_bar.setValue(100)
            QMessageBox.information(self, 'Success', 'PDF comparison with enhanced OCR completed successfully!')
            self.view_excel_button.setEnabled(True)
            
            # Send completion email notification
            try:
                subject = "PDF Comparison Completed"
                body = f"Your PDF comparison for {self.lob_name} has been completed successfully.\n\nResults have been saved to: {self.excel_path}"
                # Note: You would need to get the user's email from the login system
                # send_email("user@example.com", subject, body, [self.excel_path])
            except Exception as e:
                logging.error(f"Failed to send completion email: {e}")

    def open_excel_viewer(self):
        if self.excel_path and os.path.exists(self.excel_path):
            self.excel_viewer = ExcelViewerWindow(self.excel_path)
            self.excel_viewer.show()
        else:
            QMessageBox.warning(self, "Error", "No Excel file found. Please run comparison first.")

    def closeEvent(self, event):
        try:
            if os.path.exists('temp'):
                for file in os.listdir('temp'):
                    os.remove(os.path.join('temp', file))
                os.rmdir('temp')
        except Exception as e:
            logging.error(f"Error cleaning up temp files: {str(e)}")
        event.accept()

# ---------------------------
# WaitDialog
# ---------------------------
class WaitDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(self.windowFlags() | Qt.FramelessWindowHint)
        self.setStyleSheet(f"background-color: {COLORS['login_bg']}; border-radius: 10px;")
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        label = QLabel("Please Wait")
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet(f"color: {COLORS['login_primary']}; font-size: 16px;")
        layout.addWidget(label, alignment=Qt.AlignCenter)

        spinner_label = QLabel()
        spinner_label.setAlignment(Qt.AlignCenter)
        spinner_gif = QMovie("temp/sandglass.gif")
        if not spinner_gif.isValid():
            spinner_label.setText("⌛")
            spinner_label.setStyleSheet("font-size: 30px;")
        else:
            spinner_label.setMovie(spinner_gif)
            spinner_gif.start()

        layout.addWidget(spinner_label, alignment=Qt.AlignCenter)
        self.setFixedSize(200, 150)

# ---------------------------
# Splash Screen
# ---------------------------
class SplashScreen(QSplashScreen):
    def __init__(self):
        if os.path.exists("temp/logo.png"):
            pixmap = QPixmap("temp/logo.png")
        else:
            response = requests.get(LOGO_URL)
            os.makedirs("temp", exist_ok=True)
            with open("temp/logo.png", "wb") as f:
                f.write(response.content)
            pixmap = QPixmap("temp/logo.png")

        super().__init__(pixmap)
        self.setWindowFlag(Qt.FramelessWindowHint)
        self.movie = QMovie("temp/splash.gif")
        if not self.movie.isValid():
            self.movie = None

    def start(self):
        if self.movie:
            self.setMovie(self.movie)
            self.movie.start()
        QTimer.singleShot(3000, self.close)

# ---------------------------
# LoginWindow, SignupDialog, etc.
# ---------------------------
login_window = None

class LoginWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Marble Box - Enhanced Login")
        self.setGeometry(100, 100, 400, 300)
        if os.path.exists("temp/icon.png"):
            self.setWindowIcon(QIcon("temp/icon.png"))
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.settings = QSettings("MarbleBox", "Login")
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        logo_label = QLabel()
        if os.path.exists("temp/icon.png"):
            pixmap = QPixmap("temp/icon.png")
        else:
            response = requests.get(LOGO_URL)
            with open("temp/logo.png", "wb") as f:
                f.write(response.content)
            pixmap = QPixmap("temp/logo.png")

        logo_label.setPixmap(pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo_label)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Enter your email")
        self.email_input.setStyleSheet(f"padding: 8px; font-size: 14px; border: 1px solid {COLORS['login_disabled']};")
        stored_email = self.settings.value("rememberedEmail", "")
        if stored_email:
            self.email_input.setText(stored_email)
        layout.addWidget(self.email_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Enter your password")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setStyleSheet(f"padding: 8px; font-size: 14px; border: 1px solid {COLORS['login_disabled']};")
        layout.addWidget(self.password_input)

        self.remember_checkbox = QCheckBox("Remember me")
        self.remember_checkbox.setStyleSheet(f"font-size: 12px; color: {COLORS['login_primary']};")
        if stored_email:
            self.remember_checkbox.setChecked(True)
        layout.addWidget(self.remember_checkbox)

        login_button = QPushButton("Login")
        login_button.setStyleSheet(f"background-color: {COLORS['login_primary']}; color: white; padding: 10px;")
        login_button.setAutoDefault(True)
        login_button.setDefault(True)
        login_button.clicked.connect(self.login)
        layout.addWidget(login_button)

        forgot_button = QPushButton("Forgot Password?")
        forgot_button.setStyleSheet(f"background-color: transparent; color: {COLORS['login_alert']}; text-decoration: underline;")
        forgot_button.clicked.connect(self.forgot_password)
        layout.addWidget(forgot_button)

        signup_button = QPushButton("Sign Up")
        signup_button.setStyleSheet(f"background-color: {COLORS['login_secondary']}; color: white; padding: 10px;")
        signup_button.clicked.connect(self.open_signup)
        layout.addWidget(signup_button)

    def login(self):
        email = self.email_input.text().strip()
        password = self.password_input.text().strip()
        if not email or not password:
            QMessageBox.warning(self, "Error", "Please enter both email and password.")
            return
        users = load_users()
        if email in users and users[email]["password"] == password:
            QMessageBox.information(self, "Success", "Login successful!")
            if self.remember_checkbox.isChecked():
                self.settings.setValue("rememberedEmail", email)
            else:
                self.settings.remove("rememberedEmail")
            
            # Send login notification email
            try:
                subject = "Marble Box Login Notification"
                body = f"Hello,\n\nYou have successfully logged into the Marble Box PDF Comparison Tool.\n\nTime: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\nRegards,\nMarble Box Team"
                send_email(email, subject, body)
            except Exception as e:
                logging.error(f"Failed to send login notification: {e}")
            
            self.open_welcome_page()
        else:
            QMessageBox.critical(self, "Error", "Invalid email or password.")

    def forgot_password(self):
        email = self.email_input.text().strip()
        if not email:
            QMessageBox.warning(self, "Error", "Please enter your email for password reset.")
            return
        users = load_users()
        if email not in users:
            QMessageBox.critical(self, "Error", "Email not registered.")
            return
        temp_code = generate_temp_code()
        QMessageBox.information(self, "Reset Password", f"A reset code has been sent to your email.\nTemporary Code: {temp_code}")
        subject = "Marble Box Password Reset Code"
        body = f"Hello,\n\nYour reset code is: {temp_code}\n\nUse this code to reset your password.\n\nRegards,\nMarble Box Team"
        
        # Enhanced email sending with progress feedback
        email_thread = send_email(email, subject, body)
        
        def on_email_sent(success, error):
            if success:
                logging.info("Password reset email sent successfully")
            else:
                logging.error(f"Failed to send password reset email: {error}")
                QMessageBox.warning(self, "Email Error", f"Failed to send reset email: {error}")
        
        email_thread.finished_signal.connect(on_email_sent)
        
        dlg = ResetPasswordDialog(email, temp_code)
        dlg.exec_()

    def open_signup(self):
        dlg = SignupDialog()
        dlg.exec_()

    def open_welcome_page(self):
        msg_box = QMessageBox()
        msg_box.setWindowTitle("SharePoint Access")
        msg_box.setText("Access to SharePoint and share drive granted 😀\nEnhanced OCR and Email System Active")
        QTimer.singleShot(2000, msg_box.accept)
        msg_box.exec_()
        self.welcome_page = WelcomePage()
        self.welcome_page.show()
        self.close()

class SignupDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Sign Up - Enhanced")
        self.setGeometry(150, 150, 350, 200)
        if os.path.exists("temp/icon.png"):
            self.setWindowIcon(QIcon("temp/icon.png"))
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Enter your email")
        layout.addRow("Email:", self.email_input)

        signup_button = QPushButton("Register")
        signup_button.setStyleSheet(f"background-color: {COLORS['login_secondary']}; color: white; padding: 8px;")
        signup_button.clicked.connect(self.register)
        layout.addRow(signup_button)

    def register(self):
        email = self.email_input.text().strip()
        if not email:
            QMessageBox.warning(self, "Error", "Please enter an email address.")
            return
        users = load_users()
        if email in users:
            QMessageBox.warning(self, "Error", "Email already registered. Please log in.")
            self.close()
            return
        temp_password = generate_temp_code()
        QMessageBox.information(self, "Registration", f"A temporary password has been sent to your email.\nTemporary Password: {temp_password}")
        subject = "Marble Box Registration - Temporary Password"
        body = f"Hello,\n\nThank you for registering with Marble Box Enhanced PDF Comparison Tool.\nYour temporary password is: {temp_password}\n\nUse this to complete your registration.\n\nRegards,\nMarble Box Team"
        
        # Enhanced email sending
        email_thread = send_email(email, subject, body)
        
        def on_email_sent(success, error):
            if success:
                logging.info("Registration email sent successfully")
                dlg = CompleteRegistrationDialog(email, temp_password)
                dlg.exec_()
            else:
                logging.error(f"Failed to send registration email: {error}")
                QMessageBox.critical(self, "Email Error", f"Failed to send registration email: {error}")
        
        email_thread.finished_signal.connect(on_email_sent)
        self.close()

class CompleteRegistrationDialog(QDialog):
    def __init__(self, email, temp_password):
        super().__init__()
        self.email = email
        self.temp_password = temp_password
        self.setWindowTitle("Complete Registration - Enhanced")
        self.setGeometry(200, 200, 350, 250)
        if os.path.exists("temp/icon.png"):
            self.setWindowIcon(QIcon("temp/icon.png"))
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)
        self.temp_input = QLineEdit()
        self.temp_input.setPlaceholderText("Enter temporary password")
        layout.addRow("Temporary Password:", self.temp_input)

        self.new_password_input = QLineEdit()
        self.new_password_input.setPlaceholderText("Enter new password")
        self.new_password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("New Password:", self.new_password_input)

        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setPlaceholderText("Confirm new password")
        self.confirm_password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("Confirm Password:", self.confirm_password_input)

        complete_button = QPushButton("Complete Registration")
        complete_button.setStyleSheet(f"background-color: {COLORS['login_primary']}; color: white; padding: 8px;")
        complete_button.clicked.connect(self.complete_registration)
        layout.addRow(complete_button)

    def complete_registration(self):
        temp_entered = self.temp_input.text().strip()
        new_password = self.new_password_input.text().strip()
        confirm_password = self.confirm_password_input.text().strip()

        if temp_entered != self.temp_password:
            QMessageBox.critical(self, "Error", "Temporary password does not match.")
            return
        if not new_password or new_password != confirm_password:
            QMessageBox.critical(self, "Error", "Passwords do not match or are empty.")
            return

        users = load_users()
        users[self.email] = {"password": new_password, "verified": True}
        save_users(users)

        QMessageBox.information(self, "Success", "Registration completed. You can now log in.")
        subject = "Marble Box Registration Completed"
        body = f"Hello,\n\nYour registration for the Enhanced Marble Box PDF Comparison Tool is now complete. You can log in with your new password.\n\nRegards,\nMarble Box Team"
        
        # Enhanced email sending
        email_thread = send_email(self.email, subject, body)
        
        def on_email_sent(success, error):
            if not success:
                logging.error(f"Failed to send completion email: {error}")
        
        email_thread.finished_signal.connect(on_email_sent)
        self.close()

class ResetPasswordDialog(QDialog):
    def __init__(self, email, temp_code):
        super().__init__()
        self.email = email
        self.temp_code = temp_code
        self.setWindowTitle("Reset Password - Enhanced")
        self.setGeometry(200, 200, 350, 250)
        if os.path.exists("temp/icon.png"):
            self.setWindowIcon(QIcon("temp/icon.png"))
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)
        self.code_input = QLineEdit()
        self.code_input.setPlaceholderText("Enter reset code")
        layout.addRow("Reset Code:", self.code_input)

        self.new_password_input = QLineEdit()
        self.new_password_input.setPlaceholderText("Enter new password")
        self.new_password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("New Password:", self.new_password_input)

        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setPlaceholderText("Confirm new password")
        self.confirm_password_input.setEchoMode(QLineEdit.Password)
        layout.addRow("Confirm Password:", self.confirm_password_input)

        reset_button = QPushButton("Reset Password")
        reset_button.setStyleSheet(f"background-color: {COLORS['login_primary']}; color: white; padding: 8px;")
        reset_button.clicked.connect(self.reset_password)
        layout.addRow(reset_button)

    def reset_password(self):
        code_entered = self.code_input.text().strip()
        new_password = self.new_password_input.text().strip()
        confirm_password = self.confirm_password_input.text().strip()

        if code_entered != self.temp_code:
            QMessageBox.critical(self, "Error", "Reset code does not match.")
            return
        if not new_password or new_password != confirm_password:
            QMessageBox.critical(self, "Error", "Passwords do not match or are empty.")
            return

        users = load_users()
        if self.email in users:
            users[self.email]["password"] = new_password
            save_users(users)
            QMessageBox.information(self, "Success", "Password has been reset. A confirmation email has been sent.")
            subject = "Marble Box Password Reset Confirmation"
            body = f"Hello,\n\nYour password for the Enhanced Marble Box PDF Comparison Tool has been reset successfully.\n\nRegards,\nMarble Box Team"
            
            # Enhanced email sending
            email_thread = send_email(self.email, subject, body)
            
            def on_email_sent(success, error):
                if not success:
                    logging.error(f"Failed to send confirmation email: {error}")
            
            email_thread.finished_signal.connect(on_email_sent)
            self.close()
        else:
            QMessageBox.critical(self, "Error", "Email not found.")
            self.close()

# ---------------------------
# Enhanced Email Test Dialog
# ---------------------------
class EmailTestDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Email System Test")
        self.setGeometry(200, 200, 400, 300)
        self.setStyleSheet(f"background-color: {COLORS['login_bg']};")
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Email input
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Enter test email address")
        layout.addWidget(QLabel("Test Email Address:"))
        layout.addWidget(self.email_input)
        
        # Subject input
        self.subject_input = QLineEdit()
        self.subject_input.setText("Marble Box Email System Test")
        layout.addWidget(QLabel("Subject:"))
        layout.addWidget(self.subject_input)
        
        # Message input
        self.message_input = QLineEdit()
        self.message_input.setText("This is a test email from the enhanced Marble Box system.")
        layout.addWidget(QLabel("Message:"))
        layout.addWidget(self.message_input)
        
        # Send button
        send_button = QPushButton("Send Test Email")
        send_button.setStyleSheet(f"background-color: {COLORS['login_primary']}; color: white; padding: 10px;")
        send_button.clicked.connect(self.send_test_email)
        layout.addWidget(send_button)
        
        # Status label
        self.status_label = QLabel("Ready to send test email")
        self.status_label.setStyleSheet(f"color: {COLORS['login_primary']};")
        layout.addWidget(self.status_label)

    def send_test_email(self):
        email = self.email_input.text().strip()
        subject = self.subject_input.text().strip()
        message = self.message_input.text().strip()
        
        if not email or not subject or not message:
            QMessageBox.warning(self, "Error", "Please fill in all fields.")
            return
        
        self.status_label.setText("Sending test email...")
        
        # Send test email
        email_thread = send_email(email, subject, message)
        
        def on_email_sent(success, error):
            if success:
                self.status_label.setText("Test email sent successfully!")
                QMessageBox.information(self, "Success", "Test email sent successfully!")
            else:
                self.status_label.setText(f"Failed to send email: {error}")
                QMessageBox.critical(self, "Error", f"Failed to send email: {error}")
        
        def on_progress(message):
            self.status_label.setText(f"Progress: {message}")
        
        email_thread.finished_signal.connect(on_email_sent)
        email_thread.progress_signal.connect(on_progress)

# ---------------------------
# Main Application with Enhanced Features
# ---------------------------
login_window = None

def main():
    try:
        app = QApplication(sys.argv)
        app.setStyle("Fusion")
        
        # Create temp directory for resources
        os.makedirs("temp", exist_ok=True)
        
        # Download and cache resources
        try:
            if not os.path.exists("temp/logo.png"):
                response = requests.get(LOGO_URL)
                with open("temp/logo.png", "wb") as f:
                    f.write(response.content)
            
            if not os.path.exists("temp/icon.png"):
                response = requests.get(ICON_URL)
                with open("temp/icon.png", "wb") as f:
                    f.write(response.content)
        except Exception as e:
            logging.warning(f"Failed to download resources: {e}")
        
        # Show splash screen
        splash = SplashScreen()
        splash.show()
        splash.start()

        def on_splash_finished():
            global login_window
            login_window = LoginWindow()
            login_window.show()
            splash.close()

        QTimer.singleShot(3000, on_splash_finished)
        
        # Add menu for email testing (for debugging)
        def show_email_test():
            email_test = EmailTestDialog()
            email_test.exec_()
        
        # You can add this to the main window menu if needed
        # main_window.menuBar().addAction("Test Email", show_email_test)
        
        logging.info("Enhanced Marble Box PDF Comparison Tool started successfully")
        logging.info("Features: Enhanced OCR, Improved Email System, Better Error Handling")
        
        sys.exit(app.exec_())
        
    except Exception as e:
        logging.error(f"Application error: {str(e)}")
        if 'app' in locals():
            QMessageBox.critical(None, "Application Error", f"Failed to start application: {str(e)}")

if __name__ == '__main__':
    main()
